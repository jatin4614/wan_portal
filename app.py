"""
WAN Network Monitoring Portal
A Flask-based dashboard for monitoring 17-node WAN with OFC, alternate media,
terminal equipment, radio equipment, and fault logs.

Run: python app.py
Default URL: http://0.0.0.0:5000
"""

import os
import io
import json
import sqlite3
import hashlib
import secrets
from functools import wraps
from datetime import datetime, date, timedelta
from contextlib import closing

from flask import (
    Flask, render_template, request, jsonify, send_file,
    redirect, url_for, flash, session, abort
)
import pandas as pd
from openpyxl import Workbook

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------
APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(APP_DIR, "instance", "wan_portal.db")
MBTILES_DIR = os.path.join(APP_DIR, "instance", "tiles")
MBTILES_PATH = os.path.join(MBTILES_DIR, "active.mbtiles")
PMTILES_PATH = os.path.join(MBTILES_DIR, "active.pmtiles")

app = Flask(__name__)
app.secret_key = "change-this-in-production-please"
# Upload cap. Default 16 GB (effectively "no cap" for normal use). Override
# via environment variable. MAX_UPLOAD_MB=0 disables the cap entirely.
_max_mb = int(os.environ.get("MAX_UPLOAD_MB", "16384"))
if _max_mb > 0:
    app.config["MAX_CONTENT_LENGTH"] = _max_mb * 1024 * 1024
else:
    app.config["MAX_CONTENT_LENGTH"] = None  # no limit


# ---------------------------------------------------------------------------
# Sub-path hosting (IIS application under /wanportal, nginx location, etc.)
# ---------------------------------------------------------------------------
# When this app is mounted under a URL prefix, the reverse proxy (IIS
# HttpPlatformHandler) forwards the FULL path -- e.g. /wanportal/login -- to us.
# Flask's routes are defined at the root (/login), so without help every request
# would 404. This middleware moves the configured prefix out of PATH_INFO and
# into SCRIPT_NAME, which makes Flask's routing match AND makes url_for() /
# request.script_root generate correctly-prefixed URLs.
#
# Controlled by the URL_PREFIX environment variable (set in web.config on the
# server). Empty/unset => served at the root, so a plain `python app.py` still
# works at http://localhost:5000/ with no prefix.
class PrefixMiddleware:
    def __init__(self, wsgi_app, prefix=""):
        self.wsgi_app = wsgi_app
        p = (prefix or "").strip().strip("/")
        self.prefix = ("/" + p) if p else ""

    def __call__(self, environ, start_response):
        if self.prefix:
            path = environ.get("PATH_INFO", "")
            if path == self.prefix or path.startswith(self.prefix + "/"):
                environ["SCRIPT_NAME"] = self.prefix + environ.get("SCRIPT_NAME", "")
                environ["PATH_INFO"] = path[len(self.prefix):] or "/"
        return self.wsgi_app(environ, start_response)


app.wsgi_app = PrefixMiddleware(app.wsgi_app, os.environ.get("URL_PREFIX", ""))


# Friendly 413 handler that explains exactly what to do
@app.errorhandler(413)
def too_large(e):
    cap_bytes = app.config.get("MAX_CONTENT_LENGTH")
    cap_mb = (cap_bytes // (1024 * 1024)) if cap_bytes else "unlimited"
    msg = (f"The file is larger than the current upload cap of {cap_mb} MB. "
           f"Raise it by setting the MAX_UPLOAD_MB environment variable "
           f"(set MAX_UPLOAD_MB=0 to disable the cap completely) and "
           f"restart the server. If you are behind a reverse proxy "
           f"(nginx / Apache / IIS), also raise the proxy's upload limit.")
    if request.path.startswith("/api/"):
        return jsonify({"error": msg, "limit_mb": cap_mb}), 413
    return msg, 413


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    schema = """
    CREATE TABLE IF NOT EXISTS nodes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        latitude REAL,
        longitude REAL,
        region TEXT,
        node_type TEXT,
        formation TEXT,
        unit TEXT,
        coy TEXT,
        contact_no TEXT,
        remarks TEXT
    );

    CREATE TABLE IF NOT EXISTS ofc_links (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        link_id TEXT UNIQUE NOT NULL,
        from_node TEXT NOT NULL,
        to_node TEXT NOT NULL,
        distance_km REAL,
        year_laid INTEGER,
        no_of_fiber INTEGER,
        ribbon_count INTEGER DEFAULT 1,
        loss_db REAL,
        no_dark_fiber INTEGER,
        cable_type TEXT,
        margin_db REAL,
        status TEXT DEFAULT 'UP',
        last_trace_date TEXT,
        trace_taken_by TEXT,
        remarks TEXT,
        updated_at TEXT
    );

    CREATE TABLE IF NOT EXISTS alternate_media (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        media_type TEXT NOT NULL,
        from_node TEXT NOT NULL,
        to_node TEXT,
        spec TEXT,
        hop_distance_km REAL,
        status TEXT DEFAULT 'UP',
        remarks TEXT,
        updated_at TEXT
    );

    CREATE TABLE IF NOT EXISTS terminal_equipment (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        equipment_id TEXT UNIQUE NOT NULL,
        location TEXT,
        eqpt_type TEXT,
        eth_ports INTEGER,
        e1_voice_ports INTEGER,
        capacity TEXT,
        year_purchased INTEGER,
        status TEXT DEFAULT 'OPERATIONAL',
        last_checked_on TEXT,
        last_checked_by TEXT,
        remarks TEXT
    );

    CREATE TABLE IF NOT EXISTS radio_equipment (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        equipment_id TEXT UNIQUE NOT NULL,
        radio_type TEXT,
        location TEXT,
        frequency TEXT,
        year_purchased INTEGER,
        status TEXT DEFAULT 'OPERATIONAL',
        remarks TEXT
    );

    CREATE TABLE IF NOT EXISTS fault_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fault_id TEXT UNIQUE NOT NULL,
        category TEXT,
        affected_link TEXT,
        affected_node TEXT,
        reported_at TEXT NOT NULL,
        resolved_at TEXT,
        severity TEXT,
        description TEXT,
        action_taken TEXT,
        status TEXT DEFAULT 'OPEN'
    );

    -- Roles: 'Admin' plus one per formation (e.g. '1 Brigade')
    CREATE TABLE IF NOT EXISTS roles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        description TEXT,
        created_at TEXT
    );

    -- Users: simple username + password hash + role assignment
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL,
        full_name TEXT,
        created_at TEXT,
        last_login TEXT
    );

    -- Misc communication activities logged on the Daily Outage page
    CREATE TABLE IF NOT EXISTS misc_activities (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        content TEXT NOT NULL,
        author_username TEXT NOT NULL,
        author_role TEXT NOT NULL,
        activity_date TEXT,
        created_at TEXT,
        updated_at TEXT
    );

    -- Ribbons inside an OFC link (one OFC has multiple ribbons; each ribbon has multiple fibers)
    CREATE TABLE IF NOT EXISTS ofc_ribbons (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ofc_link_id INTEGER NOT NULL,
        ribbon_number INTEGER NOT NULL,
        fiber_count INTEGER NOT NULL,
        remarks TEXT,
        UNIQUE(ofc_link_id, ribbon_number),
        FOREIGN KEY(ofc_link_id) REFERENCES ofc_links(id) ON DELETE CASCADE
    );

    -- Individual fiber strands. Each fiber may be connected to a port on each end (from/to).
    CREATE TABLE IF NOT EXISTS ofc_fibers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ribbon_id INTEGER NOT NULL,
        fiber_number INTEGER NOT NULL,
        status TEXT DEFAULT 'FREE',
        from_port_id INTEGER,
        to_port_id INTEGER,
        remarks TEXT,
        UNIQUE(ribbon_id, fiber_number),
        FOREIGN KEY(ribbon_id) REFERENCES ofc_ribbons(id) ON DELETE CASCADE,
        FOREIGN KEY(from_port_id) REFERENCES equipment_ports(id) ON DELETE SET NULL,
        FOREIGN KEY(to_port_id) REFERENCES equipment_ports(id) ON DELETE SET NULL
    );

    -- Ports on terminal equipment.
    CREATE TABLE IF NOT EXISTS equipment_ports (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        terminal_equipment_id INTEGER NOT NULL,
        port_label TEXT NOT NULL,
        port_type TEXT,
        bandwidth TEXT,
        status TEXT DEFAULT 'UNUSED',
        remarks TEXT,
        UNIQUE(terminal_equipment_id, port_label),
        FOREIGN KEY(terminal_equipment_id) REFERENCES terminal_equipment(id) ON DELETE CASCADE
    );

    -- Reusable equipment-type templates. The port_template is a JSON list of
    -- port specs: e.g. [{"label":"Eth-1","port_type":"Optical","bandwidth":"10 Gbps","count":16},
    --                   {"label":"E1-","port_type":"Electronic","bandwidth":"2 Mbps","count":8}]
    CREATE TABLE IF NOT EXISTS equipment_types (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        description TEXT,
        port_template TEXT,
        remarks TEXT,
        created_at TEXT
    );
    """
    with closing(get_db()) as conn:
        conn.executescript(schema)
        # Add new columns to existing tables if upgrading from earlier schema
        existing_ofc_cols = [c[1] for c in conn.execute("PRAGMA table_info(ofc_links)").fetchall()]
        if "last_trace_date" not in existing_ofc_cols:
            conn.execute("ALTER TABLE ofc_links ADD COLUMN last_trace_date TEXT")
        if "trace_taken_by" not in existing_ofc_cols:
            conn.execute("ALTER TABLE ofc_links ADD COLUMN trace_taken_by TEXT")
        if "ribbon_count" not in existing_ofc_cols:
            conn.execute("ALTER TABLE ofc_links ADD COLUMN ribbon_count INTEGER DEFAULT 1")
        # Migrate nodes: add formation, unit, contact_no columns
        existing_node_cols = [c[1] for c in conn.execute("PRAGMA table_info(nodes)").fetchall()]
        for col in ("formation", "unit", "coy", "contact_no"):
            if col not in existing_node_cols:
                conn.execute(f"ALTER TABLE nodes ADD COLUMN {col} TEXT")
        # Migrate terminal_equipment: add last_checked_on, last_checked_by
        existing_eq_cols = [c[1] for c in conn.execute("PRAGMA table_info(terminal_equipment)").fetchall()]
        for col in ("last_checked_on", "last_checked_by"):
            if col not in existing_eq_cols:
                conn.execute(f"ALTER TABLE terminal_equipment ADD COLUMN {col} TEXT")
        conn.commit()


# ---------------------------------------------------------------------------
# Default equipment types (port templates)
# Each template is a list of port specs:
#   { "label_prefix": "Eth-", "port_type": "Optical", "bandwidth": "10 Gbps", "count": 16 }
# When equipment is created with that type, ports get auto-generated:
#   Eth-1, Eth-2, ... Eth-16 each Optical/10Gbps
# ---------------------------------------------------------------------------
DEFAULT_EQUIPMENT_TYPES = [
    {
        "name": "Core Router",
        "description": "High-capacity backbone router",
        "ports": [
            {"label_prefix": "Eth-", "port_type": "Optical",   "bandwidth": "100 Gbps", "count": 8},
            {"label_prefix": "GE-",  "port_type": "Optical",   "bandwidth": "10 Gbps",  "count": 16},
            {"label_prefix": "E1-",  "port_type": "Electronic","bandwidth": "2 Mbps",   "count": 8},
        ],
    },
    {
        "name": "Edge Router",
        "description": "Branch / edge router",
        "ports": [
            {"label_prefix": "Eth-", "port_type": "Optical",   "bandwidth": "10 Gbps", "count": 8},
            {"label_prefix": "GE-",  "port_type": "Optical",   "bandwidth": "1 Gbps",  "count": 8},
            {"label_prefix": "E1-",  "port_type": "Electronic","bandwidth": "2 Mbps",  "count": 4},
        ],
    },
    {
        "name": "MAR",
        "description": "Multi-Access Router",
        "ports": [
            {"label_prefix": "Eth-", "port_type": "Optical",   "bandwidth": "1 Gbps", "count": 8},
            {"label_prefix": "E1-",  "port_type": "Electronic","bandwidth": "2 Mbps", "count": 16},
        ],
    },
    {
        "name": "MER",
        "description": "Multi-service Edge Router",
        "ports": [
            {"label_prefix": "Eth-", "port_type": "Optical",   "bandwidth": "1 Gbps", "count": 4},
            {"label_prefix": "E1-",  "port_type": "Electronic","bandwidth": "2 Mbps", "count": 8},
        ],
    },
    {
        "name": "STM-16",
        "description": "Synchronous Transport Module - 16",
        "ports": [
            {"label_prefix": "STM-", "port_type": "Optical", "bandwidth": "STM-16", "count": 4},
        ],
    },
    {
        "name": "Optimux",
        "description": "Optical multiplexer",
        "ports": [
            {"label_prefix": "Opt-", "port_type": "Optical",   "bandwidth": "STM-1",  "count": 4},
            {"label_prefix": "E1-",  "port_type": "Electronic","bandwidth": "2 Mbps", "count": 8},
        ],
    },
    {
        "name": "FXO/FXS",
        "description": "Voice gateway",
        "ports": [
            {"label_prefix": "Voice-", "port_type": "Electronic", "bandwidth": "Voice", "count": 24},
        ],
    },
    {
        "name": "L2 Switch",
        "description": "Layer-2 switch (Coy-level)",
        "ports": [
            {"label_prefix": "GE-",  "port_type": "Optical",   "bandwidth": "1 Gbps",   "count": 8},
            {"label_prefix": "FE-",  "port_type": "Electronic","bandwidth": "100 Mbps", "count": 16},
        ],
    },
    {
        "name": "OTE",
        "description": "Optical Terminal Equipment (Post-level)",
        "ports": [
            {"label_prefix": "Opt-", "port_type": "Optical",   "bandwidth": "1 Gbps",   "count": 2},
            {"label_prefix": "Eth-", "port_type": "Electronic","bandwidth": "100 Mbps", "count": 4},
        ],
    },
]


def seed_default_equipment_types():
    """Seed only if the equipment_types table is empty."""
    with closing(get_db()) as conn:
        n = conn.execute("SELECT COUNT(*) FROM equipment_types").fetchone()[0]
        if n > 0:
            return
        now = datetime.now().isoformat(timespec="seconds")
        for et in DEFAULT_EQUIPMENT_TYPES:
            conn.execute(
                "INSERT INTO equipment_types (name, description, port_template, created_at) VALUES (?,?,?,?)",
                (et["name"], et["description"], json.dumps(et["ports"]), now),
            )
        conn.commit()


# ---------------------------------------------------------------------------
# Auth: helpers, decorators, default Admin seeding
# ---------------------------------------------------------------------------
def hash_password(plain):
    """Salted SHA-256. Good enough for an internal LAN tool with no PII."""
    salt = secrets.token_hex(8)
    h = hashlib.sha256((salt + plain).encode("utf-8")).hexdigest()
    return f"{salt}${h}"


def verify_password(stored, plain):
    try:
        salt, h = stored.split("$", 1)
    except ValueError:
        return False
    return hashlib.sha256((salt + plain).encode("utf-8")).hexdigest() == h


def seed_default_admin():
    """Make sure 'Admin' role exists and a default admin user is present."""
    with closing(get_db()) as conn:
        now = datetime.now().isoformat(timespec="seconds")
        # Admin role
        if not conn.execute("SELECT 1 FROM roles WHERE name='Admin'").fetchone():
            conn.execute(
                "INSERT INTO roles (name, description, created_at) VALUES (?,?,?)",
                ("Admin", "Full system access — sees and edits everything", now),
            )
        # Default admin user (only if no users exist at all)
        if conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
            conn.execute(
                "INSERT INTO users (username, password_hash, role, full_name, created_at) VALUES (?,?,?,?,?)",
                ("admin", hash_password("admin"), "Admin", "Default Administrator", now),
            )
        conn.commit()


def sync_roles_with_formations():
    """For each distinct formation name in nodes, make sure a role exists."""
    with closing(get_db()) as conn:
        formations = conn.execute(
            "SELECT DISTINCT formation FROM nodes WHERE formation IS NOT NULL AND formation != ''"
        ).fetchall()
        now = datetime.now().isoformat(timespec="seconds")
        for f in formations:
            fname = f["formation"]
            exists = conn.execute("SELECT 1 FROM roles WHERE name=?", (fname,)).fetchone()
            if not exists:
                conn.execute(
                    "INSERT INTO roles (name, description, created_at) VALUES (?,?,?)",
                    (fname, f"Auto-created for formation '{fname}'", now),
                )
        conn.commit()


def current_user():
    """Return the dict {username, role} for the logged-in user, or None."""
    uname = session.get("username")
    role = session.get("role")
    if not uname or not role:
        return None
    return {"username": uname, "role": role}


def is_admin():
    u = current_user()
    return bool(u and u["role"] == "Admin")


def user_formation_scope():
    """Return the formation name a non-Admin user is restricted to, or None for Admin / no user."""
    u = current_user()
    if not u or u["role"] == "Admin":
        return None
    return u["role"]  # role name == formation name for non-Admin users


def scope_filter(table):
    """Return (extra_sql, params) appended to a WHERE for the given table.
    Returns ('', ()) for Admin/no-user (no filtering applied).
    For non-Admin: limit rows to those linked to the user's formation."""
    formation = user_formation_scope()
    if not formation:
        return "", ()
    if table == "nodes":
        return "AND formation = ?", (formation,)
    if table in ("ofc_links", "alternate_media"):
        # Match if either endpoint's node has this formation
        return (
            "AND (from_node IN (SELECT name FROM nodes WHERE formation=?) "
            " OR  to_node   IN (SELECT name FROM nodes WHERE formation=?))",
            (formation, formation),
        )
    if table in ("terminal_equipment", "radio_equipment"):
        return (
            "AND location IN (SELECT name FROM nodes WHERE formation=?)",
            (formation,),
        )
    if table == "fault_logs":
        # affected_node directly or affected_link endpoint
        return (
            "AND ("
            "  affected_node IN (SELECT name FROM nodes WHERE formation=?) "
            "  OR affected_link IN ("
            "      SELECT link_id FROM ofc_links "
            "      WHERE from_node IN (SELECT name FROM nodes WHERE formation=?) "
            "         OR to_node   IN (SELECT name FROM nodes WHERE formation=?)"
            "  )"
            ")",
            (formation, formation, formation),
        )
    return "", ()


def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not current_user():
            # API requests get JSON, page requests get a redirect
            if request.path.startswith("/api/"):
                return jsonify({"error": "auth required"}), 401
            # Carry the mount prefix in `next` so post-login redirect stays
            # inside the app when hosted under a sub-path (script_root is '' at root).
            return redirect(url_for("login_view", next=request.script_root + request.path))
        return f(*args, **kwargs)
    return wrapper


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not current_user():
            if request.path.startswith("/api/"):
                return jsonify({"error": "auth required"}), 401
            return redirect(url_for("login_view"))
        if not is_admin():
            if request.path.startswith("/api/"):
                return jsonify({"error": "admin only"}), 403
            abort(403)
        return f(*args, **kwargs)
    return wrapper


@app.context_processor
def inject_user():
    """Make current_user available inside Jinja templates."""
    return {
        "current_user": current_user(),
        "is_admin": is_admin(),
    }


# ---------------------------------------------------------------------------
# Mock data seeding (17 nodes, realistic Indian WAN-style topology)
# ---------------------------------------------------------------------------
def seed_mock_data():
    """Seed only if the DB is empty."""
    with closing(get_db()) as conn:
        n = conn.execute("SELECT COUNT(*) FROM nodes").fetchone()[0]
        if n > 0:
            return

        # 17 nodes spread across India with realistic lat/long
        nodes = [
            ("HQ-DELHI",      28.6139, 77.2090, "North",   "Core HQ"),
            ("MUMBAI",        19.0760, 72.8777, "West",    "Regional HQ"),
            ("KOLKATA",       22.5726, 88.3639, "East",    "Regional HQ"),
            ("CHENNAI",       13.0827, 80.2707, "South",   "Regional HQ"),
            ("BANGALORE",     12.9716, 77.5946, "South",   "Edge"),
            ("HYDERABAD",     17.3850, 78.4867, "South",   "Edge"),
            ("PUNE",          18.5204, 73.8567, "West",    "Edge"),
            ("AHMEDABAD",     23.0225, 72.5714, "West",    "Edge"),
            ("JAIPUR",        26.9124, 75.7873, "North",   "Edge"),
            ("LUCKNOW",       26.8467, 80.9462, "North",   "Edge"),
            ("BHOPAL",        23.2599, 77.4126, "Central", "Edge"),
            ("NAGPUR",        21.1458, 79.0882, "Central", "Edge"),
            ("PATNA",         25.5941, 85.1376, "East",    "Edge"),
            ("BHUBANESWAR",   20.2961, 85.8245, "East",    "Edge"),
            ("GUWAHATI",      26.1445, 91.7362, "NE",      "Edge"),
            ("SRINAGAR",      34.0837, 74.7973, "North",   "Edge"),
            ("CHANDIGARH",    30.7333, 76.7794, "North",   "Edge"),
        ]
        conn.executemany(
            "INSERT INTO nodes (name, latitude, longitude, region, node_type) VALUES (?,?,?,?,?)",
            nodes,
        )

        # OFC links - realistic mesh with redundancy
        # tuple: link_id, from, to, dist, year, fibers, ribbons, loss, dark, cable, margin, status, last_trace_date, trace_taken_by, remarks
        ofc = [
            ("OFC-001", "HQ-DELHI",   "JAIPUR",     280, 2018, 24, 2, 0.22, 6,  "G.652D",  6.5, "UP",       "2025-03-15", "Team Alpha",  "Primary backbone"),
            ("OFC-002", "HQ-DELHI",   "LUCKNOW",    495, 2017, 24, 2, 0.24, 4,  "G.652D",  5.8, "UP",       "2025-02-22", "Team Bravo",  ""),
            ("OFC-003", "HQ-DELHI",   "CHANDIGARH", 245, 2019, 48, 4, 0.20, 12, "G.652D",  7.2, "UP",       "2025-04-10", "Team Alpha",  ""),
            ("OFC-004", "HQ-DELHI",   "MUMBAI",    1420, 2016, 48, 4, 0.28, 8,  "G.652D",  4.5, "UP",       "2025-01-30", "Team Charlie", "Long-haul DWDM"),
            ("OFC-005", "MUMBAI",     "PUNE",       150, 2020, 24, 2, 0.18, 6,  "G.652D",  8.1, "UP",       "2025-03-28", "Team Delta",   ""),
            ("OFC-006", "MUMBAI",     "AHMEDABAD",  525, 2018, 24, 2, 0.25, 4,  "G.652D",  5.5, "UP",       "2025-02-15", "Team Charlie", ""),
            ("OFC-007", "MUMBAI",     "NAGPUR",     820, 2017, 24, 2, 0.30, 2,  "G.652D",  3.2, "DOWN",     "2024-11-12", "Team Charlie", "Cable cut at KM 412"),
            ("OFC-008", "PUNE",       "BANGALORE",  840, 2019, 24, 2, 0.26, 4,  "G.652D",  4.8, "UP",       "2025-03-05", "Team Delta",   ""),
            ("OFC-009", "BANGALORE",  "CHENNAI",    345, 2018, 48, 4, 0.21, 10, "G.652D",  6.9, "UP",       "2025-04-02", "Team Echo",    ""),
            ("OFC-010", "BANGALORE",  "HYDERABAD",  570, 2019, 24, 2, 0.24, 6,  "G.652D",  5.6, "UP",       "2025-03-18", "Team Echo",    ""),
            ("OFC-011", "CHENNAI",    "HYDERABAD",  625, 2016, 24, 2, 0.32, 2,  "G.652B",  2.8, "DEGRADED", "2025-01-08", "Team Echo",    "High loss, replace planned"),
            ("OFC-012", "HYDERABAD",  "NAGPUR",     500, 2020, 24, 2, 0.19, 8,  "G.652D",  7.5, "UP",       "2025-04-12", "Team Foxtrot", ""),
            ("OFC-013", "NAGPUR",     "BHOPAL",     355, 2019, 24, 2, 0.22, 6,  "G.652D",  6.8, "UP",       "2025-03-22", "Team Foxtrot", ""),
            ("OFC-014", "BHOPAL",     "JAIPUR",     595, 2018, 24, 2, 0.25, 4,  "G.652D",  5.4, "UP",       "2025-02-28", "Team Alpha",   ""),
            ("OFC-015", "LUCKNOW",    "PATNA",      535, 2017, 24, 2, 0.27, 4,  "G.652D",  4.7, "UP",       "2025-02-05", "Team Bravo",   ""),
            ("OFC-016", "PATNA",      "KOLKATA",    580, 2018, 24, 2, 0.24, 6,  "G.652D",  5.9, "UP",       "2025-03-10", "Team Bravo",   ""),
            ("OFC-017", "KOLKATA",    "BHUBANESWAR",440, 2019, 48, 4, 0.21, 12, "G.652D",  7.1, "UP",       "2025-03-25", "Team Golf",    ""),
            ("OFC-018", "BHUBANESWAR","CHENNAI",    975, 2016, 24, 2, 0.30, 2,  "G.652D",  3.5, "UP",       "2025-01-15", "Team Echo",    ""),
            ("OFC-019", "KOLKATA",    "GUWAHATI",   985, 2017, 24, 2, 0.28, 4,  "G.652D",  4.2, "UP",       "2025-02-18", "Team Golf",    ""),
            ("OFC-020", "HQ-DELHI",   "SRINAGAR",   875, 2015, 24, 2, 0.35, 0,  "G.652B",  1.5, "DEGRADED", "2024-12-20", "Team Alpha",   "Aging cable, no spare fiber"),
            ("OFC-021", "CHANDIGARH", "SRINAGAR",   650, 2019, 24, 2, 0.26, 6,  "G.652D",  5.1, "UP",       "2025-03-30", "Team Alpha",   ""),
            ("OFC-022", "JAIPUR",     "AHMEDABAD",  625, 2020, 24, 2, 0.20, 8,  "G.652D",  7.3, "UP",       "2025-04-05", "Team Charlie", ""),
        ]
        now = datetime.now().isoformat(timespec="seconds")
        conn.executemany(
            """INSERT INTO ofc_links
               (link_id, from_node, to_node, distance_km, year_laid, no_of_fiber, ribbon_count,
                loss_db, no_dark_fiber, cable_type, margin_db, status,
                last_trace_date, trace_taken_by, remarks, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            [(*r, now) for r in ofc],
        )

        # Alternate media (some Satellite/URRF/RRF entries have no to_node — single-location)
        alt = [
            ("Microwave",       "HQ-DELHI",   "JAIPUR",     "7 GHz / 155 Mbps",  282, "UP",       ""),
            ("Microwave",       "MUMBAI",     "PUNE",       "11 GHz / 622 Mbps", 152, "UP",       ""),
            ("Satellite",       "HQ-DELHI",   None,         "Ku-Band / 10 Mbps", None, "UP",      "VSAT terminal"),
            ("Satellite",       "SRINAGAR",   None,         "Ku-Band / 8 Mbps",  None, "UP",      "Backup only"),
            ("Satellite",       "GUWAHATI",   None,         "Ku-Band / 8 Mbps",  None, "UP",      ""),
            ("Broadband Radio", "BANGALORE",  "CHENNAI",    "5 GHz / 300 Mbps",  348, "UP",       ""),
            ("Broadband Radio", "PUNE",       "BANGALORE",  "5 GHz / 300 Mbps",  842, "DOWN",     "Tower power issue"),
            ("BTS",             "HQ-DELHI",   "CHANDIGARH", "4G LTE",            245, "UP",       ""),
            ("Copper JFC",      "MUMBAI",     "AHMEDABAD",  "E1 / 2 Mbps",       520, "UP",       "Legacy backup"),
            ("Cellular Tower",  "JAIPUR",     "AHMEDABAD",  "4G / 100 Mbps",     625, "UP",       ""),
            ("Microwave",       "BHOPAL",     "NAGPUR",     "7 GHz / 155 Mbps",  358, "DEGRADED", "Rain fade observed"),
            ("Microwave",       "PATNA",      "LUCKNOW",    "7 GHz / 155 Mbps",  535, "UP",       ""),
            ("URRF",            "HQ-DELHI",   "CHANDIGARH", "UHF / 50W",         245, "UP",       "Primary URRF link"),
            ("URRF",            "BHOPAL",     "NAGPUR",     "UHF / 30W",         358, "UP",       ""),
            ("RRF",             "LUCKNOW",    "PATNA",      "VHF / 25W",         535, "UP",       ""),
            ("RRF",             "CHENNAI",    "HYDERABAD",  "VHF / 25W",         625, "DEGRADED", "Antenna alignment off"),
        ]
        conn.executemany(
            """INSERT INTO alternate_media
               (media_type, from_node, to_node, spec, hop_distance_km, status, remarks, updated_at)
               VALUES (?,?,?,?,?,?,?,?)""",
            [(*r, now) for r in alt],
        )

        # Terminal equipment
        eqpt_types = ["Core Router", "Edge Router", "MAR", "MER", "STM-16", "Optimux", "FXO/FXS"]
        terminal = []
        for i, node in enumerate([n[0] for n in nodes]):
            terminal.extend([
                (f"CR-{i+1:02d}",  node, "Core Router", 24, 8,  "100 Gbps", 2020 + (i % 4),  "OPERATIONAL", ""),
                (f"ER-{i+1:02d}",  node, "Edge Router", 16, 4,  "10 Gbps",  2019 + (i % 5),  "OPERATIONAL", ""),
                (f"MAR-{i+1:02d}", node, "MAR",         8,  16, "1 Gbps",   2018 + (i % 5),  "OPERATIONAL", ""),
                (f"OPT-{i+1:02d}", node, "Optimux",     4,  8,  "STM-1",    2015 + (i % 7),  "OPERATIONAL" if i % 5 else "FAULTY", ""),
            ])
        conn.executemany(
            """INSERT INTO terminal_equipment
               (equipment_id, location, eqpt_type, eth_ports, e1_voice_ports,
                capacity, year_purchased, status, remarks)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            terminal,
        )

        # Radio equipment
        radio = []
        radio_types = ["Combat Network Radio", "Motorola", "Mobile Comn", "ISAT Phone"]
        for i, node in enumerate([n[0] for n in nodes]):
            rt = radio_types[i % 4]
            radio.append((f"RAD-{i+1:02d}", rt, node, f"{400 + i*5} MHz", 2018 + (i % 6),
                         "OPERATIONAL" if i % 7 else "FAULTY", ""))
        conn.executemany(
            """INSERT INTO radio_equipment
               (equipment_id, radio_type, location, frequency, year_purchased, status, remarks)
               VALUES (?,?,?,?,?,?,?)""",
            radio,
        )

        # Fault logs - mix of open and resolved
        today = datetime.now()
        faults = [
            ("FLT-2025-001", "OFC",       "OFC-007", "MUMBAI/NAGPUR",
             (today - timedelta(days=2, hours=4)).isoformat(timespec="seconds"), None,
             "HIGH",     "Cable cut at KM 412 due to road work",     "Splice team dispatched", "OPEN"),
            ("FLT-2025-002", "OFC",       "OFC-011", "CHENNAI/HYDERABAD",
             (today - timedelta(days=10)).isoformat(timespec="seconds"),
             (today - timedelta(days=9)).isoformat(timespec="seconds"),
             "MEDIUM",   "Loss exceeded threshold",                  "Connector cleaned, monitoring", "RESOLVED"),
            ("FLT-2025-003", "Radio",     None,      "BHOPAL",
             (today - timedelta(days=1)).isoformat(timespec="seconds"), None,
             "MEDIUM",   "Combat radio antenna alignment off",       "Tech visit scheduled", "OPEN"),
            ("FLT-2025-004", "Alt Media", None,      "PUNE/BANGALORE",
             (today - timedelta(hours=8)).isoformat(timespec="seconds"), None,
             "HIGH",     "BBR tower battery failure",                "Generator on site", "OPEN"),
            ("FLT-2025-005", "Equipment", None,      "GUWAHATI",
             (today - timedelta(days=15)).isoformat(timespec="seconds"),
             (today - timedelta(days=14, hours=6)).isoformat(timespec="seconds"),
             "LOW",      "Optimux port flapping",                    "Card replaced", "RESOLVED"),
            ("FLT-2025-006", "OFC",       "OFC-020", "DELHI/SRINAGAR",
             (today - timedelta(days=5)).isoformat(timespec="seconds"), None,
             "HIGH",     "High loss, fiber aging",                   "Replacement project initiated", "OPEN"),
            ("FLT-2025-007", "Equipment", None,      "JAIPUR",
             (today - timedelta(days=20)).isoformat(timespec="seconds"),
             (today - timedelta(days=19)).isoformat(timespec="seconds"),
             "MEDIUM",   "Edge router CPU spike",                    "Software patched", "RESOLVED"),
            ("FLT-2025-008", "Alt Media", None,      "BHOPAL/NAGPUR",
             (today - timedelta(days=3)).isoformat(timespec="seconds"), None,
             "LOW",      "MW link rain fade",                        "Monitoring weather",     "OPEN"),
        ]
        conn.executemany(
            """INSERT INTO fault_logs
               (fault_id, category, affected_link, affected_node, reported_at,
                resolved_at, severity, description, action_taken, status)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            faults,
        )

        # ----- Equipment ports -----------------------------------------------
        # Generate ports for each piece of terminal equipment based on its eth_ports + e1_voice_ports counts.
        # All ports start UNUSED. The fiber-allocation step below will mark them ACTIVE/DOWN as needed.
        all_eq = conn.execute("SELECT id, equipment_id, eqpt_type, eth_ports, e1_voice_ports, capacity, status FROM terminal_equipment").fetchall()
        port_rows = []
        for eq in all_eq:
            eth = eq["eth_ports"] or 0
            e1 = eq["e1_voice_ports"] or 0
            base_bw = eq["capacity"] or "1 Gbps"
            for p in range(1, eth + 1):
                # First quarter of ports get the equipment's base capacity; rest get 1 Gbps
                bw = base_bw if p <= max(1, eth // 4) else "1 Gbps"
                port_rows.append((eq["id"], f"Eth-{p}", "Optical", bw, "UNUSED", ""))
            for p in range(1, e1 + 1):
                port_rows.append((eq["id"], f"E1-{p}", "Electronic", "2 Mbps", "UNUSED", ""))
        conn.executemany(
            """INSERT INTO equipment_ports
               (terminal_equipment_id, port_label, port_type, bandwidth, status, remarks)
               VALUES (?,?,?,?,?,?)""",
            port_rows,
        )

        # ----- OFC ribbons + fibers ------------------------------------------
        all_ofc = conn.execute("SELECT id, link_id, from_node, to_node, no_of_fiber, ribbon_count, status FROM ofc_links").fetchall()
        for ol in all_ofc:
            n_fibers = ol["no_of_fiber"] or 0
            n_ribbons = max(1, ol["ribbon_count"] or 1)
            # Distribute fibers across ribbons as evenly as possible
            base = n_fibers // n_ribbons
            extra = n_fibers % n_ribbons
            ribbon_ids = []
            for rb in range(1, n_ribbons + 1):
                cnt = base + (1 if rb <= extra else 0)
                cur = conn.execute(
                    "INSERT INTO ofc_ribbons (ofc_link_id, ribbon_number, fiber_count) VALUES (?,?,?)",
                    (ol["id"], rb, cnt),
                )
                ribbon_ids.append((cur.lastrowid, cnt))
            # Find the equipment+ports we can use to terminate fibers at each end.
            # Pick all unused Ethernet ports at FROM and TO node respectively.
            from_ports = conn.execute(
                """SELECT p.id, p.port_label FROM equipment_ports p
                   JOIN terminal_equipment t ON p.terminal_equipment_id = t.id
                   WHERE t.location = ? AND p.port_type = 'Optical' AND p.status = 'UNUSED'
                   ORDER BY t.id, p.id""", (ol["from_node"],)
            ).fetchall()
            to_ports = conn.execute(
                """SELECT p.id, p.port_label FROM equipment_ports p
                   JOIN terminal_equipment t ON p.terminal_equipment_id = t.id
                   WHERE t.location = ? AND p.port_type = 'Optical' AND p.status = 'UNUSED'
                   ORDER BY t.id, p.id""", (ol["to_node"],)
            ).fetchall()
            from_port_iter = iter(from_ports)
            to_port_iter = iter(to_ports)

            link_status = ol["status"]
            for ribbon_id, fcount in ribbon_ids:
                for fnum in range(1, fcount + 1):
                    # Connection logic:
                    # Roughly 70% connected. 20% reserved. 10% free.
                    # If both ends have available ports we can connect.
                    bucket = (ol["id"] * 17 + ribbon_id * 11 + fnum) % 10
                    fp, tp = None, None
                    fstatus = "FREE"
                    if bucket < 7:
                        fp_row = next(from_port_iter, None)
                        tp_row = next(to_port_iter, None)
                        if fp_row and tp_row:
                            fp, tp = fp_row["id"], tp_row["id"]
                            # If link is DOWN -> all connected fibers become CONNECTED-DOWN
                            # If link DEGRADED -> ~30% of connected fibers go CONNECTED-DOWN
                            if link_status == "DOWN":
                                fstatus = "CONNECTED-DOWN"
                            elif link_status == "DEGRADED" and bucket < 2:
                                fstatus = "CONNECTED-DOWN"
                            else:
                                fstatus = "CONNECTED-ACTIVE"
                    elif bucket < 9:
                        fstatus = "RESERVED"
                    else:
                        fstatus = "FREE"
                    conn.execute(
                        """INSERT INTO ofc_fibers
                           (ribbon_id, fiber_number, status, from_port_id, to_port_id)
                           VALUES (?,?,?,?,?)""",
                        (ribbon_id, fnum, fstatus, fp, tp),
                    )
                    # Mark used ports as ACTIVE/DOWN to stay consistent
                    if fp:
                        conn.execute("UPDATE equipment_ports SET status=? WHERE id=?",
                                     ("ACTIVE" if fstatus == "CONNECTED-ACTIVE" else "DOWN", fp))
                    if tp:
                        conn.execute("UPDATE equipment_ports SET status=? WHERE id=?",
                                     ("ACTIVE" if fstatus == "CONNECTED-ACTIVE" else "DOWN", tp))

        # Mark some ports DOWN on FAULTY equipment (every other Eth port)
        faulty_eq = conn.execute("SELECT id FROM terminal_equipment WHERE status != 'OPERATIONAL'").fetchall()
        for eq in faulty_eq:
            ports = conn.execute(
                "SELECT id FROM equipment_ports WHERE terminal_equipment_id=? AND port_type='Optical' ORDER BY id",
                (eq["id"],)
            ).fetchall()
            for i, p in enumerate(ports):
                if i % 2 == 0:
                    conn.execute("UPDATE equipment_ports SET status='DOWN' WHERE id=?", (p["id"],))

        conn.commit()


# ---------------------------------------------------------------------------
# Hierarchical formation seed: 1 Div + 1 Div-Access + 5 Bde + 5 Bde-Access
# + 20 Bn + 40 Coy + 100 Post = 172 nodes
# ---------------------------------------------------------------------------
NODE_TYPE_OPTIONS = ["Div", "Div-Access", "Bde", "Bde-Access", "Bn", "Coy", "Post"]


def seed_hierarchy_data(clear_first=True):
    """Generate a realistic 172-node formation hierarchy in North India with
    OFC links following the chain of command. Returns counts dict."""
    import random
    rng = random.Random(42)

    with closing(get_db()) as conn:
        if clear_first:
            # Wipe in dependency order
            for t in ("ofc_fibers", "ofc_ribbons", "ofc_links",
                      "alternate_media", "fault_logs",
                      "equipment_ports", "terminal_equipment",
                      "radio_equipment", "nodes"):
                conn.execute(f"DELETE FROM {t}")
        now = datetime.now().isoformat(timespec="seconds")

        # Helper: jitter coords within a radius (degrees)
        def jit(lat, lng, radius_deg):
            return (lat + rng.uniform(-radius_deg, radius_deg),
                    lng + rng.uniform(-radius_deg, radius_deg))

        nodes_to_insert = []  # tuples of (name, lat, lng, region, node_type, formation, unit, coy, contact_no, remarks)
        ofc_to_insert = []    # tuples for OFC links

        # Helper: add a node row
        def add_node(name, lat, lng, region, ntype, formation, unit, contact, remark="", coy=None):
            nodes_to_insert.append((name, round(lat, 4), round(lng, 4), region,
                                    ntype, formation, unit, coy, contact, remark))

        # ---- Division HQ (Udhampur is a real Div HQ in N. India) ----
        div_lat, div_lng = 32.9159, 75.1416  # Udhampur
        div_name = "DIV-HQ"
        add_node(div_name, div_lat, div_lng, "J&K", "Div",
                 "Northern Command Div", "Div HQ", "+91-1992-200001", "Divisional HQ")

        # ---- Div-Access (co-located near Div) ----
        dla, dln = jit(div_lat, div_lng, 0.05)
        add_node("DIV-ACCESS", dla, dln, "J&K", "Div-Access",
                 "Northern Command Div", "Div Access Sig", "+91-1992-200002",
                 "Division access node")

        # OFC link Div ↔ Div-Access
        ofc_to_insert.append(("OFC-DIV-001", "DIV-HQ", "DIV-ACCESS", 5, 2022, 24, 2,
                              0.18, 8, "G.652D", 8.0, "UP",
                              "2025-04-20", "Sig Det Alpha", "Div redundant access"))

        # ---- 5 Brigades around the Division (within ~150km / 1.5° radius) ----
        bde_centers = []
        bde_seeds = [
            ("BDE-1", "Jammu",     32.7266, 74.8570),
            ("BDE-2", "Srinagar",  34.0837, 74.7973),
            ("BDE-3", "Leh",       34.1526, 77.5771),
            ("BDE-4", "Pathankot", 32.2746, 75.6521),
            ("BDE-5", "Akhnoor",   32.8987, 74.7290),
        ]
        for i, (bname, region, blat, blng) in enumerate(bde_seeds, 1):
            formation = f"{i} Brigade"
            add_node(bname, blat, blng, region, "Bde", formation, "Bde HQ Sig",
                     f"+91-1991-30{i:02d}01", f"{i} Bde HQ at {region}")
            bde_centers.append((bname, blat, blng, region, formation))
            # OFC: Div ↔ Bde
            ofc_to_insert.append((f"OFC-DIV-BDE-{i:02d}",
                                  "DIV-HQ", bname,
                                  rng.randint(80, 220), 2018 + (i % 4), 24, 2,
                                  round(rng.uniform(0.20, 0.30), 2),
                                  rng.randint(2, 6), "G.652D",
                                  round(rng.uniform(4.5, 7.5), 1),
                                  "UP", "2025-03-15", f"Sig Det B{i}",
                                  f"Div trunk to {bname}"))
            # Bde-Access node
            ala, aln = jit(blat, blng, 0.05)
            access_name = f"BDE-{i}-ACCESS"
            add_node(access_name, ala, aln, region, "Bde-Access", formation,
                     "Bde Access Sig", f"+91-1991-30{i:02d}02",
                     f"Access node co-located with {bname}")
            ofc_to_insert.append((f"OFC-BDE-ACC-{i:02d}", bname, access_name,
                                  4, 2022, 12, 1, 0.18, 4, "G.652D", 7.5,
                                  "UP", "2025-04-10", f"Sig Det B{i}",
                                  "Bde access redundant"))

        # ---- 20 Battalions (4 per Bde) ----
        bn_count = 0
        bn_centers = []
        for bname, blat, blng, region, formation in bde_centers:
            bde_idx = int(bname.split("-")[1])
            for k in range(1, 5):
                bn_count += 1
                nlat, nlng = jit(blat, blng, 0.4)  # ~50 km
                bn_name = f"BN-{bn_count:02d}"
                bn_unit = f"{bn_count} Battalion"
                add_node(bn_name, nlat, nlng, region, "Bn", formation, bn_unit,
                         f"+91-1991-40{bn_count:02d}01",
                         f"{bn_unit} under {formation}")
                bn_centers.append((bn_name, nlat, nlng, region, formation, bn_unit))
                # OFC: Bde ↔ Bn
                ofc_to_insert.append((f"OFC-BDE{bde_idx:02d}-BN{bn_count:02d}",
                                      bname, bn_name,
                                      rng.randint(20, 70), 2019 + (k % 4), 12, 1,
                                      round(rng.uniform(0.22, 0.32), 2),
                                      rng.randint(0, 3), "G.652D",
                                      round(rng.uniform(3.5, 6.0), 1),
                                      "UP" if rng.random() > 0.05 else "DEGRADED",
                                      "2025-02-20", f"Sig Det B{bde_idx}",
                                      ""))

        # ---- 40 Companies (2 per Bn) ----
        coy_count = 0
        coy_centers = []
        for bn_name, blat, blng, region, formation, bn_unit in bn_centers:
            for k in range(1, 3):
                coy_count += 1
                clat, clng = jit(blat, blng, 0.15)  # ~20 km
                coy_name = f"COY-{coy_count:03d}"
                coy_label = f"Coy {chr(64+k)}"  # "Coy A", "Coy B"
                coy_unit_str = f"{bn_unit} - {coy_label}"  # display string still kept for ofc remarks
                # For Coy nodes: unit = parent Bn's unit; coy = own Coy designation
                add_node(coy_name, clat, clng, region, "Coy", formation, bn_unit,
                         f"+91-1991-50{coy_count:03d}",
                         coy_unit_str, coy=coy_label)
                coy_centers.append((coy_name, clat, clng, region, formation, bn_unit, coy_label))
                # OFC: Bn ↔ Coy
                ofc_to_insert.append((f"OFC-BN-COY-{coy_count:03d}",
                                      bn_name, coy_name,
                                      rng.randint(8, 30), 2020 + (k % 3), 6, 1,
                                      round(rng.uniform(0.24, 0.36), 2),
                                      rng.randint(0, 2), "G.652D",
                                      round(rng.uniform(2.5, 5.0), 1),
                                      "UP" if rng.random() > 0.08 else "DEGRADED",
                                      "2025-01-15", f"Sig Det B{int(bn_name.split('-')[1]) % 5 + 1}",
                                      ""))

        # ---- 100 Posts (~2-3 per Coy, distributed) ----
        post_count = 0
        for idx, (coy_name, clat, clng, region, formation, bn_unit, coy_label) in enumerate(coy_centers):
            posts_here = 3 if idx < 20 else 2
            for k in range(1, posts_here + 1):
                post_count += 1
                if post_count > 100:
                    break
                plat, plng = jit(clat, clng, 0.08)  # ~10 km
                post_name = f"POST-{post_count:03d}"
                # For Post nodes: unit = parent Bn's unit; coy = parent Coy designation
                add_node(post_name, plat, plng, region, "Post", formation, bn_unit,
                         f"+91-1991-60{post_count:03d}",
                         f"Forward post under {bn_unit} {coy_label}",
                         coy=coy_label)
                # OFC: Coy ↔ Post (some are degraded/down for realism)
                status = "UP"
                if rng.random() < 0.06: status = "DEGRADED"
                if rng.random() < 0.03: status = "DOWN"
                ofc_to_insert.append((f"OFC-COY-POST-{post_count:03d}",
                                      coy_name, post_name,
                                      rng.randint(3, 15), 2021 + (k % 2), 4, 1,
                                      round(rng.uniform(0.26, 0.40), 2),
                                      rng.randint(0, 1), "G.652D",
                                      round(rng.uniform(1.5, 3.5), 1),
                                      status, "2024-12-10", f"Sig Det B{int(coy_name.split('-')[1]) % 5 + 1}",
                                      ""))

        # Bulk insert
        conn.executemany(
            """INSERT INTO nodes (name, latitude, longitude, region, node_type,
                                  formation, unit, coy, contact_no, remarks)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            nodes_to_insert
        )
        conn.executemany(
            """INSERT INTO ofc_links
               (link_id, from_node, to_node, distance_km, year_laid, no_of_fiber, ribbon_count,
                loss_db, no_dark_fiber, cable_type, margin_db, status,
                last_trace_date, trace_taken_by, remarks, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            [(*r, now) for r in ofc_to_insert]
        )

        # Auto-generate ribbons + fibers for each new OFC
        for ol in conn.execute("SELECT id, no_of_fiber, ribbon_count FROM ofc_links").fetchall():
            n_fibers = ol["no_of_fiber"] or 0
            n_ribbons = max(1, ol["ribbon_count"] or 1)
            base = n_fibers // n_ribbons
            extra = n_fibers % n_ribbons
            for rb in range(1, n_ribbons + 1):
                cnt = base + (1 if rb <= extra else 0)
                cur = conn.execute(
                    "INSERT INTO ofc_ribbons (ofc_link_id, ribbon_number, fiber_count) VALUES (?,?,?)",
                    (ol["id"], rb, cnt),
                )
                rid = cur.lastrowid
                for fnum in range(1, cnt + 1):
                    conn.execute(
                        "INSERT INTO ofc_fibers (ribbon_id, fiber_number, status) VALUES (?,?,?)",
                        (rid, fnum, "FREE"),
                    )

        # ---- Terminal equipment per node, with auto-generated ports ----
        # Mapping per node type:
        #   Div, Div-Access  → Core Router OR Edge Router
        #   Bde, Bde-Access  → Edge Router
        #   Bn               → MER
        #   Coy              → L2 Switch
        #   Post             → OTE
        type_to_eqpt = {
            "Div":         ["Core Router", "Edge Router"],   # one of each
            "Div-Access":  ["Edge Router"],
            "Bde":         ["Edge Router"],
            "Bde-Access":  ["Edge Router"],
            "Bn":          ["MER"],
            "Coy":         ["L2 Switch"],
            "Post":        ["OTE"],
        }

        # Pre-fetch templates (id, name, port_template_json)
        type_lookup = {
            r["name"]: json.loads(r["port_template"]) if r["port_template"] else []
            for r in conn.execute("SELECT name, port_template FROM equipment_types").fetchall()
        }

        eq_count = 0
        port_count = 0
        # Re-fetch the freshly inserted nodes so we have their ids/names paired
        all_nodes = conn.execute("SELECT id, name, node_type FROM nodes").fetchall()
        for node in all_nodes:
            eqpt_types = type_to_eqpt.get(node["node_type"], [])
            for idx, eqpt_name in enumerate(eqpt_types, 1):
                eq_count += 1
                # Use full node name to guarantee uniqueness across the whole hierarchy
                eq_id = f"EQ-{node['name']}-{idx:02d}"
                # Light realistic variation
                year = 2018 + (eq_count % 6)
                # ~5% faulty
                status = "FAULTY" if (eq_count % 19) == 0 else "OPERATIONAL"
                cur = conn.execute(
                    """INSERT INTO terminal_equipment
                       (equipment_id, location, eqpt_type, year_purchased, status,
                        last_checked_on, last_checked_by, remarks)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (eq_id, node["name"], eqpt_name, year, status,
                     "2025-04-15",
                     f"Sig Det B{(eq_count % 5) + 1}",
                     ""),
                )
                new_eq_id = cur.lastrowid
                # Auto-create ports from template
                for spec in type_lookup.get(eqpt_name, []):
                    prefix = spec.get("label_prefix", "Port-")
                    ptype = spec.get("port_type", "Optical")
                    bw = spec.get("bandwidth", "1 Gbps")
                    cnt = int(spec.get("count", 0) or 0)
                    for p in range(1, cnt + 1):
                        port_count += 1
                        try:
                            conn.execute(
                                """INSERT INTO equipment_ports
                                   (terminal_equipment_id, port_label, port_type, bandwidth, status)
                                   VALUES (?,?,?,?,?)""",
                                (new_eq_id, f"{prefix}{p}", ptype, bw, "UNUSED"),
                            )
                        except Exception:
                            pass

        # ---- Alternate Media ------------------------------------------------
        # For each parent-child link in the formation chain, add a backup wireless
        # path. Plus a Satellite per Bde and URRF/RRF for some Posts.
        alt_count = 0

        # Helper: insert an alternate media row
        def add_alt(media_type, from_n, to_n, spec, hop_km, status="UP", remark=""):
            nonlocal alt_count
            alt_count += 1
            conn.execute(
                """INSERT INTO alternate_media
                   (media_type, from_node, to_node, spec, hop_distance_km, status, remarks, updated_at)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (media_type, from_n, to_n, spec, hop_km, status, remark, now),
            )

        # Div ↔ Div-Access: redundant Microwave
        add_alt("Microwave", "DIV-HQ", "DIV-ACCESS",
                "7 GHz · 1 Gbps", 5, "UP", "Backup for OFC-DIV-001")

        # Div ↔ each Bde: Microwave
        for bname, blat, blng, region, formation in bde_centers:
            add_alt("Microwave", "DIV-HQ", bname,
                    f"7 GHz · {rng.choice([300, 600, 1000])} Mbps",
                    rng.randint(80, 220),
                    "UP" if rng.random() > 0.1 else "DEGRADED",
                    f"MW backup to {bname}")

        # One Satellite per Bde (single-location, no to_node)
        for bname, blat, blng, region, formation in bde_centers:
            add_alt("Satellite", bname, None,
                    "VSAT Ku-band · 4 Mbps", None,
                    "UP" if rng.random() > 0.05 else "DEGRADED",
                    f"VSAT terminal at {bname}")

        # Bde ↔ Bn: Microwave or Broadband Radio (mixed)
        for bn_name, blat, blng, region, formation, bn_unit in bn_centers:
            # Find parent Bde from formation
            parent_bde = next((b[0] for b in bde_centers if b[4] == formation), None)
            if not parent_bde:
                continue
            mtype = rng.choice(["Microwave", "Broadband Radio"])
            spec = "11 GHz · 200 Mbps" if mtype == "Microwave" else "5 GHz · 100 Mbps"
            add_alt(mtype, parent_bde, bn_name, spec,
                    rng.randint(15, 50),
                    "UP" if rng.random() > 0.12 else "DEGRADED")

        # Bn ↔ Coy: Broadband Radio
        for coy_name, clat, clng, region, formation, bn_unit, coy_label in coy_centers:
            # Find the actual Bn name from bn_centers by matching bn_unit
            parent_bn_name = next((b[0] for b in bn_centers if b[5] == bn_unit), None)
            if not parent_bn_name:
                continue
            add_alt("Broadband Radio", parent_bn_name, coy_name,
                    "5 GHz · 50 Mbps", rng.randint(5, 25),
                    "UP" if rng.random() > 0.1 else "DEGRADED")

        # Coy ↔ Post: BTS / Cellular Tower (alternates) for the first 60 posts
        post_rows_for_alt = conn.execute(
            "SELECT name, region FROM nodes WHERE node_type='Post' ORDER BY name"
        ).fetchall()
        for idx, post in enumerate(post_rows_for_alt[:60]):
            # Find the parent Coy from the post's OFC link (Coy ↔ Post)
            parent = conn.execute(
                "SELECT from_node FROM ofc_links WHERE to_node=? AND link_id LIKE 'OFC-COY-POST-%'",
                (post["name"],)
            ).fetchone()
            if parent:
                mtype = "BTS" if idx % 2 == 0 else "Cellular Tower"
                add_alt(mtype, parent["from_node"], post["name"],
                        "Cell · 4G · 25 Mbps", rng.randint(2, 12),
                        "UP" if rng.random() > 0.15 else "DEGRADED")

        # URRF/RRF single-location at 30 forward posts
        for post in post_rows_for_alt[60:90]:
            mtype = rng.choice(["URRF", "RRF"])
            add_alt(mtype, post["name"], None,
                    "VHF · 25 W", None,
                    "UP" if rng.random() > 0.1 else "DEGRADED",
                    "Forward post emergency backup")

        # Copper JFC links at remaining 10 posts (last-mile copper backup)
        for post in post_rows_for_alt[90:100]:
            parent = conn.execute(
                "SELECT from_node FROM ofc_links WHERE to_node=? AND link_id LIKE 'OFC-COY-POST-%'",
                (post["name"],)
            ).fetchone()
            if parent:
                add_alt("Copper JFC", parent["from_node"], post["name"],
                        "Copper Twisted Pair · 2 Mbps",
                        rng.randint(1, 5),
                        "UP" if rng.random() > 0.18 else "DEGRADED",
                        "Copper JFC backup link")

        # ---- Radio Equipment -------------------------------------------------
        # Per node-type radio loadout:
        #   Div, Div-Access  : Combat Network Radio + ISAT Phone
        #   Bde, Bde-Access  : Combat Network Radio + ISAT Phone
        #   Bn               : Combat Network Radio + Motorola
        #   Coy              : Combat Network Radio + Mobile Comn
        #   Post             : Combat Network Radio (sole)
        radio_loadout = {
            "Div":         ["Combat Network Radio", "ISAT Phone"],
            "Div-Access":  ["Combat Network Radio", "ISAT Phone"],
            "Bde":         ["Combat Network Radio", "ISAT Phone"],
            "Bde-Access":  ["Combat Network Radio", "ISAT Phone"],
            "Bn":          ["Combat Network Radio", "Motorola"],
            "Coy":         ["Combat Network Radio", "Mobile Comn"],
            "Post":        ["Combat Network Radio"],
        }
        radio_freq = {
            "Combat Network Radio": "VHF 30-88 MHz",
            "ISAT Phone":           "L-Band 1.5 GHz",
            "Motorola":             "UHF 400-470 MHz",
            "Mobile Comn":          "GSM 900/1800",
        }
        radio_count = 0
        for node in all_nodes:
            loadout = radio_loadout.get(node["node_type"], [])
            for idx, radio_name in enumerate(loadout, 1):
                radio_count += 1
                rid = f"RAD-{node['name']}-{idx:02d}"
                year = 2018 + (radio_count % 7)
                status = "FAULTY" if (radio_count % 23) == 0 else "OPERATIONAL"
                conn.execute(
                    """INSERT INTO radio_equipment
                       (equipment_id, radio_type, location, frequency, year_purchased, status, remarks)
                       VALUES (?,?,?,?,?,?,?)""",
                    (rid, radio_name, node["name"], radio_freq.get(radio_name, ""), year, status, ""),
                )

        conn.commit()
        return {
            "nodes": len(nodes_to_insert),
            "ofc_links": len(ofc_to_insert),
            "terminal_equipment": eq_count,
            "ports": port_count,
            "alternate_media": alt_count,
            "radio_equipment": radio_count,
        }


@app.route("/api/seed-hierarchy", methods=["POST"])
@admin_required
def seed_hierarchy_endpoint():
    counts = seed_hierarchy_data(clear_first=True)
    # Auto-create roles for all newly-discovered formations
    sync_roles_with_formations()
    return jsonify({"ok": True, **counts})


# ---------------------------------------------------------------------------
# KPI / aggregation functions
# ---------------------------------------------------------------------------
def compute_kpis():
    with closing(get_db()) as conn:
        ofc_extra,  ofc_p  = scope_filter("ofc_links")
        alt_extra,  alt_p  = scope_filter("alternate_media")
        eq_extra,   eq_p   = scope_filter("terminal_equipment")
        rad_extra,  rad_p  = scope_filter("radio_equipment")
        flt_extra,  flt_p  = scope_filter("fault_logs")

        ofc = conn.execute(
            "SELECT status, distance_km FROM ofc_links WHERE 1=1 " + ofc_extra, ofc_p
        ).fetchall()
        alt = conn.execute(
            "SELECT status, media_type, hop_distance_km FROM alternate_media WHERE 1=1 " + alt_extra, alt_p
        ).fetchall()
        eq = conn.execute("SELECT status FROM terminal_equipment WHERE 1=1 " + eq_extra, eq_p).fetchall()
        rad = conn.execute("SELECT status FROM radio_equipment WHERE 1=1 " + rad_extra, rad_p).fetchall()
        faults = conn.execute("SELECT status, severity FROM fault_logs WHERE 1=1 " + flt_extra, flt_p).fetchall()

        ofc_up = sum(1 for r in ofc if r["status"] == "UP")
        ofc_down = sum(1 for r in ofc if r["status"] == "DOWN")
        ofc_deg = sum(1 for r in ofc if r["status"] == "DEGRADED")
        ofc_len = round(sum((r["distance_km"] or 0) for r in ofc), 1)

        alt_up = sum(1 for r in alt if r["status"] == "UP")
        alt_down = sum(1 for r in alt if r["status"] == "DOWN")

        def media_stats(mtype):
            rows = [r for r in alt if (r["media_type"] or "") == mtype]
            total = len(rows)
            oper = sum(1 for r in rows if r["status"] == "UP")
            length = round(sum((r["hop_distance_km"] or 0) for r in rows), 1)
            return {"total": total, "operational": oper, "length_km": length}

        jfc = media_stats("Copper JFC")
        mw = media_stats("Microwave")
        sat = media_stats("Satellite")
        bbr = media_stats("Broadband Radio")

        eq_ok = sum(1 for r in eq if r["status"] == "OPERATIONAL")
        eq_bad = sum(1 for r in eq if r["status"] != "OPERATIONAL")

        rad_ok = sum(1 for r in rad if r["status"] == "OPERATIONAL")
        rad_bad = sum(1 for r in rad if r["status"] != "OPERATIONAL")

        f_open = sum(1 for r in faults if r["status"] == "OPEN")
        f_high = sum(1 for r in faults if r["status"] == "OPEN" and r["severity"] == "HIGH")

        total_links = len(ofc)
        avail = round((ofc_up / total_links * 100), 1) if total_links else 0

        return {
            "ofc_total": total_links,
            "ofc_up": ofc_up, "ofc_down": ofc_down, "ofc_degraded": ofc_deg,
            "ofc_length_km": ofc_len,
            "jfc_total": jfc["total"], "jfc_operational": jfc["operational"],
            "jfc_length_km": jfc["length_km"],
            "mw_total": mw["total"], "mw_operational": mw["operational"],
            "sat_total": sat["total"], "sat_operational": sat["operational"],
            "bbr_total": bbr["total"], "bbr_operational": bbr["operational"],
            "alt_total": len(alt), "alt_up": alt_up, "alt_down": alt_down,
            "eq_total": len(eq), "eq_ok": eq_ok, "eq_bad": eq_bad,
            "rad_total": len(rad), "rad_ok": rad_ok, "rad_bad": rad_bad,
            "faults_open": f_open, "faults_high": f_high,
            "availability_pct": avail,
        }


def compute_node_states():
    """For each node, classify connectivity.
       Per requirement: if ANY link or terminal equipment at the node is DOWN/FAULTY,
       the node should show as ISOLATED (red).
    """
    with closing(get_db()) as conn:
        node_extra, node_p = scope_filter("nodes")
        nodes = [r["name"] for r in conn.execute(
            "SELECT name FROM nodes WHERE 1=1 " + node_extra, node_p
        ).fetchall()]
        states = {}
        for n in nodes:
            ofc_up = conn.execute(
                """SELECT COUNT(*) FROM ofc_links
                   WHERE (from_node=? OR to_node=?) AND status='UP'""", (n, n)
            ).fetchone()[0]
            ofc_deg = conn.execute(
                """SELECT COUNT(*) FROM ofc_links
                   WHERE (from_node=? OR to_node=?) AND status='DEGRADED'""", (n, n)
            ).fetchone()[0]
            ofc_down = conn.execute(
                """SELECT COUNT(*) FROM ofc_links
                   WHERE (from_node=? OR to_node=?) AND status='DOWN'""", (n, n)
            ).fetchone()[0]
            alt_up = conn.execute(
                """SELECT COUNT(*) FROM alternate_media
                   WHERE (from_node=? OR to_node=?) AND status='UP'""", (n, n)
            ).fetchone()[0]
            alt_down = conn.execute(
                """SELECT COUNT(*) FROM alternate_media
                   WHERE (from_node=? OR to_node=?) AND status='DOWN'""", (n, n)
            ).fetchone()[0]
            eq_down = conn.execute(
                """SELECT COUNT(*) FROM terminal_equipment
                   WHERE location=? AND status != 'OPERATIONAL'""", (n,)
            ).fetchone()[0]

            # Any down resource at this node => isolated/red
            if ofc_down > 0 or alt_down > 0 or eq_down > 0:
                state = "ISOLATED"
            elif ofc_up >= 2 or (ofc_up >= 1 and alt_up >= 1):
                state = "FULL"
            elif ofc_up >= 1 or alt_up >= 1 or ofc_deg >= 1:
                state = "DEGRADED"
            else:
                state = "ISOLATED"

            states[n] = {
                "state": state,
                "ofc_up": ofc_up, "ofc_deg": ofc_deg, "ofc_down": ofc_down,
                "alt_up": alt_up, "alt_down": alt_down,
                "eq_down": eq_down,
            }
        return states


def mttr_stats():
    """Compute MTTR (mean time to repair) from resolved faults."""
    with closing(get_db()) as conn:
        rows = conn.execute(
            """SELECT reported_at, resolved_at FROM fault_logs
               WHERE status='RESOLVED' AND resolved_at IS NOT NULL"""
        ).fetchall()
        if not rows:
            return {"mttr_hours": 0, "resolved_count": 0}
        deltas = []
        for r in rows:
            try:
                a = datetime.fromisoformat(r["reported_at"])
                b = datetime.fromisoformat(r["resolved_at"])
                deltas.append((b - a).total_seconds() / 3600.0)
            except Exception:
                continue
        return {
            "mttr_hours": round(sum(deltas) / len(deltas), 2) if deltas else 0,
            "resolved_count": len(deltas),
        }


# ---------------------------------------------------------------------------
# Routes - pages
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Offline tile serving from MBTiles
# ---------------------------------------------------------------------------
def _mbtiles_conn():
    if not os.path.exists(MBTILES_PATH):
        return None
    c = sqlite3.connect(MBTILES_PATH)
    c.row_factory = sqlite3.Row
    return c


@app.route("/tiles/<int:z>/<int:x>/<int:y>.png")
def serve_tile(z, x, y):
    """Serve a single tile from the active MBTiles file. MBTiles uses TMS
    coordinates (origin bottom-left) so we flip Y from the standard XYZ scheme."""
    conn = _mbtiles_conn()
    if conn is None:
        return ("MBTiles not configured", 404)
    try:
        tms_y = (1 << z) - 1 - y
        row = conn.execute(
            "SELECT tile_data FROM tiles WHERE zoom_level=? AND tile_column=? AND tile_row=?",
            (z, x, tms_y),
        ).fetchone()
        if row is None:
            return ("", 204)
        # MBTiles tiles are commonly PNG or JPEG - send as PNG by default
        from flask import Response
        resp = Response(row["tile_data"], mimetype="image/png")
        resp.headers["Cache-Control"] = "public, max-age=86400"
        return resp
    finally:
        conn.close()


@app.route("/api/tiles/info")
@login_required
def tiles_info():
    """Return metadata about the currently-loaded offline map.
    Supports both .mbtiles (raster, SQLite) and .pmtiles (vector or raster,
    single binary file)."""
    # PMTiles takes precedence — newer format, better-rendered
    if os.path.exists(PMTILES_PATH):
        try:
            info = _pmtiles_info(PMTILES_PATH)
            info["loaded"] = True
            info["format_container"] = "pmtiles"
            info["size_bytes"] = os.path.getsize(PMTILES_PATH)
            return jsonify(info)
        except Exception as e:
            return jsonify({"loaded": False, "error": f"PMTiles read failed: {e}"})

    if not os.path.exists(MBTILES_PATH):
        return jsonify({"loaded": False})
    conn = _mbtiles_conn()
    try:
        meta = {r["name"]: r["value"] for r in conn.execute("SELECT name, value FROM metadata").fetchall()}
        size = os.path.getsize(MBTILES_PATH)
        out = {
            "loaded": True,
            "format_container": "mbtiles",
            "size_bytes": size,
            "name": meta.get("name") or "Custom Map",
            "format": meta.get("format", "png"),
            "bounds": meta.get("bounds"),
            "center": meta.get("center"),
            "minzoom": int(meta.get("minzoom", 0)),
            "maxzoom": int(meta.get("maxzoom", 18)),
            "attribution": meta.get("attribution", ""),
        }
        return jsonify(out)
    except Exception as e:
        return jsonify({"loaded": False, "error": str(e)})
    finally:
        conn.close()


def _pmtiles_info(path):
    """Read the PMTiles header to extract metadata. PMTiles spec v3:
       https://github.com/protomaps/PMTiles/blob/main/spec/v3/spec.md
    The first 127 bytes are the fixed header; bytes 0-7 are the magic 'PMTiles\\x03'."""
    import struct
    with open(path, "rb") as f:
        header = f.read(127)
    if len(header) < 127 or header[:7] != b"PMTiles":
        raise ValueError("Not a valid PMTiles file (bad magic)")
    spec_version = header[7]
    if spec_version not in (2, 3):
        raise ValueError(f"Unsupported PMTiles spec version {spec_version}")

    # Spec v3 header layout (offsets):
    #   8:    root_dir offset (uint64 LE)
    #   16:   root_dir length (uint64 LE)
    #   24:   json metadata offset
    #   32:   json metadata length
    #   40:   leaf_dirs offset
    #   48:   leaf_dirs length
    #   56:   tile_data offset
    #   64:   tile_data length
    #   72:   addressed_tiles (uint64)
    #   80:   tile_entries
    #   88:   tile_contents
    #   96:   clustered (uint8)
    #   97:   internal_compression (uint8)
    #   98:   tile_compression (uint8)
    #   99:   tile_type (uint8) 0=unknown, 1=mvt, 2=png, 3=jpeg, 4=webp, 5=avif
    #   100:  min_zoom (uint8)
    #   101:  max_zoom (uint8)
    #   102:  min_lon_e7 (int32)
    #   106:  min_lat_e7 (int32)
    #   110:  max_lon_e7 (int32)
    #   114:  max_lat_e7 (int32)
    #   118:  center_zoom (uint8)
    #   119:  center_lon_e7 (int32)
    #   123:  center_lat_e7 (int32)
    json_off, json_len = struct.unpack_from("<QQ", header, 24)
    tile_type = header[99]
    min_zoom = header[100]
    max_zoom = header[101]
    min_lon = struct.unpack_from("<i", header, 102)[0] / 1e7
    min_lat = struct.unpack_from("<i", header, 106)[0] / 1e7
    max_lon = struct.unpack_from("<i", header, 110)[0] / 1e7
    max_lat = struct.unpack_from("<i", header, 114)[0] / 1e7
    center_zoom = header[118]
    center_lon = struct.unpack_from("<i", header, 119)[0] / 1e7
    center_lat = struct.unpack_from("<i", header, 123)[0] / 1e7

    type_map = {0: "unknown", 1: "mvt", 2: "png", 3: "jpeg", 4: "webp", 5: "avif"}
    internal_compression = header[97]
    # Read JSON metadata if present (may be gzipped)
    name = "Custom PMTiles Map"
    if json_len > 0:
        with open(path, "rb") as f:
            f.seek(json_off)
            blob = f.read(json_len)
        try:
            if internal_compression == 2:  # gzip
                import gzip
                blob = gzip.decompress(blob)
            meta_json = json.loads(blob.decode("utf-8"))
            name = meta_json.get("name") or name
        except Exception:
            pass

    return {
        "name": name,
        "format": type_map.get(tile_type, "unknown"),
        "is_vector": tile_type == 1,
        "minzoom": min_zoom,
        "maxzoom": max_zoom,
        "bounds": f"{min_lon},{min_lat},{max_lon},{max_lat}",
        "center": f"{center_lon},{center_lat},{center_zoom}",
        "attribution": "",
    }


@app.route("/api/pmtiles")
@login_required
def serve_pmtiles():
    """Serve the active .pmtiles file with HTTP Range support.
    The PMTiles JS library issues range requests for header + directory + tile blobs."""
    if not os.path.exists(PMTILES_PATH):
        return ("PMTiles not configured", 404)
    # send_file honors the Range header when conditional=True
    resp = send_file(
        PMTILES_PATH,
        mimetype="application/octet-stream",
        conditional=True,  # enables 206 Partial Content responses
    )
    # Some browsers / libs require Accept-Ranges to be explicit
    resp.headers["Accept-Ranges"] = "bytes"
    resp.headers["Cache-Control"] = "public, max-age=3600"
    return resp


@app.route("/api/tiles/upload", methods=["POST"])
@admin_required
def tiles_upload():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "no file"}), 400
    fname = (f.filename or "").lower()
    os.makedirs(MBTILES_DIR, exist_ok=True)

    if fname.endswith(".pmtiles"):
        # Save to PMTiles slot, then validate the header
        f.save(PMTILES_PATH)
        try:
            info = _pmtiles_info(PMTILES_PATH)
        except Exception as e:
            try:
                os.remove(PMTILES_PATH)
            except Exception:
                pass
            return jsonify({"error": f"Invalid PMTiles: {e}"}), 400
        # Clear any pre-existing MBTiles so info endpoint reports PMTiles
        if os.path.exists(MBTILES_PATH):
            try:
                os.remove(MBTILES_PATH)
            except Exception:
                pass
        return jsonify({"ok": True, "format_container": "pmtiles", **info})

    if fname.endswith(".mbtiles"):
        f.save(MBTILES_PATH)
        # Validate it's a real mbtiles
        try:
            conn = _mbtiles_conn()
            if conn is None:
                return jsonify({"error": "Could not open uploaded file"}), 400
            try:
                conn.execute("SELECT 1 FROM tiles LIMIT 1").fetchone()
            finally:
                conn.close()
        except Exception as e:
            os.remove(MBTILES_PATH)
            return jsonify({"error": f"Invalid MBTiles: {e}"}), 400
        # Clear any pre-existing PMTiles
        if os.path.exists(PMTILES_PATH):
            try:
                os.remove(PMTILES_PATH)
            except Exception:
                pass
        return jsonify({"ok": True, "format_container": "mbtiles"})

    return jsonify({"error": "Expected a .mbtiles or .pmtiles file"}), 400


@app.route("/api/tiles/delete", methods=["POST"])
@admin_required
def tiles_delete():
    removed = []
    for p, label in [(MBTILES_PATH, "mbtiles"), (PMTILES_PATH, "pmtiles")]:
        if os.path.exists(p):
            try:
                os.remove(p)
                removed.append(label)
            except Exception:
                pass
    return jsonify({"ok": True, "removed": removed})


# ---------------------------------------------------------------------------
# Equipment Types — reusable port templates
# ---------------------------------------------------------------------------
@app.route("/api/equipment-types", methods=["GET"])
def api_list_equipment_types():
    with closing(get_db()) as conn:
        rows = conn.execute("SELECT * FROM equipment_types ORDER BY name").fetchall()
        out = []
        for r in rows:
            d = dict(r)
            try:
                d["port_template"] = json.loads(d["port_template"]) if d["port_template"] else []
            except Exception:
                d["port_template"] = []
            out.append(d)
    return jsonify(out)


@app.route("/api/equipment-types", methods=["POST"])
def api_create_equipment_type():
    data = request.get_json() or {}
    if not data.get("name"):
        return jsonify({"error": "name required"}), 400
    template = data.get("port_template") or []
    try:
        with closing(get_db()) as conn:
            cur = conn.execute(
                "INSERT INTO equipment_types (name, description, port_template, remarks, created_at) VALUES (?,?,?,?,?)",
                (data["name"], data.get("description", ""), json.dumps(template),
                 data.get("remarks", ""), datetime.now().isoformat(timespec="seconds")),
            )
            conn.commit()
            return jsonify({"ok": True, "id": cur.lastrowid})
    except sqlite3.IntegrityError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/equipment-types/<int:type_id>", methods=["PUT", "DELETE"])
def api_modify_equipment_type(type_id):
    if request.method == "DELETE":
        with closing(get_db()) as conn:
            conn.execute("DELETE FROM equipment_types WHERE id=?", (type_id,))
            conn.commit()
        return jsonify({"ok": True})
    data = request.get_json() or {}
    template = data.get("port_template")
    sets, vals = [], []
    if "name" in data:
        sets.append("name=?"); vals.append(data["name"])
    if "description" in data:
        sets.append("description=?"); vals.append(data["description"])
    if template is not None:
        sets.append("port_template=?"); vals.append(json.dumps(template))
    if "remarks" in data:
        sets.append("remarks=?"); vals.append(data["remarks"])
    if not sets:
        return jsonify({"error": "no fields"}), 400
    vals.append(type_id)
    with closing(get_db()) as conn:
        conn.execute(f"UPDATE equipment_types SET {', '.join(sets)} WHERE id=?", vals)
        conn.commit()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Sample Nodes Excel — for first-time setup
# ---------------------------------------------------------------------------
@app.route("/api/sample-nodes-excel")
@admin_required
def sample_nodes_excel():
    """Generate a tiny example Nodes-only Excel file for first-time users."""
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Nodes")
    headers = ["name", "latitude", "longitude", "region", "node_type", "remarks"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    sample_rows = [
        ("HQ-DELHI",    28.6139, 77.2090, "Delhi",       "Core HQ",     "Headquarters"),
        ("CHANDIGARH",  30.7333, 76.7794, "Punjab",      "Regional HQ", ""),
        ("JAIPUR",      26.9124, 75.7873, "Rajasthan",   "Edge",        ""),
        ("SRINAGAR",    34.0837, 74.7973, "J&K",         "Edge",        ""),
        ("LUCKNOW",     26.8467, 80.9462, "UP",          "Edge",        ""),
        ("DEHRADUN",    30.3165, 78.0322, "Uttarakhand", "Edge",        ""),
    ]
    for r_idx, row in enumerate(sample_rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="sample_nodes.xlsx",
    )


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login_view():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        with closing(get_db()) as conn:
            user = conn.execute(
                "SELECT * FROM users WHERE username=?", (username,)
            ).fetchone()
        if not user or not verify_password(user["password_hash"], password):
            flash("Invalid username or password", "error")
            return redirect(url_for("login_view"))
        # Set session
        session["username"] = user["username"]
        session["role"] = user["role"]
        # Update last_login
        with closing(get_db()) as conn:
            conn.execute(
                "UPDATE users SET last_login=? WHERE id=?",
                (datetime.now().isoformat(timespec="seconds"), user["id"]),
            )
            conn.commit()
        # Only follow `next` if it is a same-origin, in-app relative path. Reject
        # absolute URLs ("http(s)://evil"), protocol-relative ("//evil") and their
        # backslash-normalized bypasses ("/\evil") to avoid an open redirect.
        next_url = request.args.get("next") or ""
        _n = next_url.replace("\\", "/")
        if not next_url or not _n.startswith("/") or _n.startswith("//"):
            next_url = url_for("dashboard")
        return redirect(next_url)
    return render_template("login.html")


@app.route("/logout")
def logout_view():
    session.clear()
    return redirect(url_for("login_view"))


@app.route("/api/users", methods=["GET"])
@admin_required
def api_list_users():
    with closing(get_db()) as conn:
        rows = conn.execute(
            "SELECT id, username, role, full_name, created_at, last_login FROM users ORDER BY username"
        ).fetchall()
    return jsonify(rows_to_list(rows))


@app.route("/api/users", methods=["POST"])
@admin_required
def api_create_user():
    data = request.get_json() or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    role = data.get("role") or ""
    full_name = data.get("full_name") or ""
    if not username or not password or not role:
        return jsonify({"error": "username, password, and role are required"}), 400
    with closing(get_db()) as conn:
        # Validate role exists
        if not conn.execute("SELECT 1 FROM roles WHERE name=?", (role,)).fetchone():
            return jsonify({"error": f"role '{role}' does not exist"}), 400
        try:
            conn.execute(
                "INSERT INTO users (username, password_hash, role, full_name, created_at) VALUES (?,?,?,?,?)",
                (username, hash_password(password), role, full_name,
                 datetime.now().isoformat(timespec="seconds")),
            )
            conn.commit()
        except sqlite3.IntegrityError:
            return jsonify({"error": "username already exists"}), 400
    return jsonify({"ok": True})


@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@admin_required
def api_delete_user(user_id):
    with closing(get_db()) as conn:
        u = conn.execute("SELECT username FROM users WHERE id=?", (user_id,)).fetchone()
        if u and u["username"] == session.get("username"):
            return jsonify({"error": "cannot delete the currently logged-in user"}), 400
        conn.execute("DELETE FROM users WHERE id=?", (user_id,))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/users/<int:user_id>/password", methods=["PUT"])
@admin_required
def api_change_user_password(user_id):
    data = request.get_json() or {}
    new_pw = data.get("password") or ""
    if len(new_pw) < 4:
        return jsonify({"error": "password too short"}), 400
    with closing(get_db()) as conn:
        conn.execute(
            "UPDATE users SET password_hash=? WHERE id=?",
            (hash_password(new_pw), user_id),
        )
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/roles", methods=["GET"])
@login_required
def api_list_roles():
    # Make sure formation-based roles are in sync first
    sync_roles_with_formations()
    with closing(get_db()) as conn:
        rows = conn.execute("SELECT * FROM roles ORDER BY name").fetchall()
    return jsonify(rows_to_list(rows))


@app.route("/users")
@admin_required
def users_view():
    return render_template("users.html")


@app.route("/")
@login_required
def dashboard():
    """Dashboard merges topology + node states + recent faults."""
    flt_extra, flt_p = scope_filter("fault_logs")
    with closing(get_db()) as conn:
        recent_faults = conn.execute(
            "SELECT * FROM fault_logs WHERE 1=1 " + flt_extra +
            " ORDER BY reported_at DESC LIMIT 8", flt_p
        ).fetchall()
    states = compute_node_states()
    state_summary = {"FULL": 0, "DEGRADED": 0, "ISOLATED": 0}
    for v in states.values():
        state_summary[v["state"]] += 1
    kpis = compute_kpis()
    mttr = mttr_stats()
    return render_template("dashboard.html",
                           kpis=kpis, mttr=mttr,
                           state_summary=state_summary,
                           recent_faults=recent_faults,
                           today=date.today().strftime("%d %b %Y"))


# Keep /topology as alias to dashboard for backward compatibility
@app.route("/topology")
@login_required
def topology():
    return redirect(url_for("dashboard"))


@app.route("/links")
@login_required
def links_view():
    return render_template("links.html")


@app.route("/nodes")
@login_required
def nodes_view():
    return render_template("nodes.html")


@app.route("/equipment")
@login_required
def equipment_view():
    return render_template("equipment.html")


@app.route("/faults")
@login_required
def faults_view():
    return render_template("faults.html")


@app.route("/analytics")
@login_required
def analytics_view():
    return render_template("analytics.html")


@app.route("/import-export")
@admin_required
def import_export_view():
    cap = app.config.get("MAX_CONTENT_LENGTH")
    return render_template(
        "import_export.html",
        upload_cap_mb=(cap // (1024 * 1024)) if cap else "unlimited",
    )


@app.route("/daily-report")
@login_required
def daily_report():
    return render_template("daily_report.html",
                           generated_at=datetime.now().strftime("%d %b %Y %H:%M"))


# ---------------------------------------------------------------------------
# API routes - data fetch
# ---------------------------------------------------------------------------
def rows_to_list(rows):
    return [dict(r) for r in rows]


# ---------------------------------------------------------------------------
# Auto-fault creation: insert a fault_logs row when something fails.
# Called from PUT handlers when a status transition into a "bad" state happens.
# ---------------------------------------------------------------------------
def _next_fault_id(conn):
    """Generate the next FLT-NNNN id by scanning existing ones."""
    row = conn.execute(
        "SELECT fault_id FROM fault_logs WHERE fault_id LIKE 'FLT-%' ORDER BY id DESC LIMIT 1"
    ).fetchone()
    if not row:
        return "FLT-0001"
    try:
        n = int(row["fault_id"].split("-")[1]) + 1
    except Exception:
        n = 1
    return f"FLT-{n:04d}"


def auto_create_fault(conn, *, category, severity, description,
                      affected_link=None, affected_node=None):
    """Insert a fault row if one isn't already open for the same item.
    De-dup: skip if there is already an OPEN fault with the same affected_link/node
    AND the same description (avoids spam when the same status PUT is repeated)."""
    # Dedup check
    existing = conn.execute(
        """SELECT id FROM fault_logs
           WHERE status='OPEN'
             AND COALESCE(affected_link,'') = COALESCE(?, '')
             AND COALESCE(affected_node,'') = COALESCE(?, '')
             AND description = ?
           LIMIT 1""",
        (affected_link, affected_node, description),
    ).fetchone()
    if existing:
        return None

    fid = _next_fault_id(conn)
    now = datetime.now().isoformat(timespec="seconds")
    conn.execute(
        """INSERT INTO fault_logs
           (fault_id, category, affected_link, affected_node, reported_at,
            severity, description, status)
           VALUES (?,?,?,?,?,?,?,?)""",
        (fid, category, affected_link, affected_node, now, severity,
         description, "OPEN"),
    )
    return fid


def _status_is_bad(table, status):
    """Return True if the given status counts as a 'fault' for the table."""
    s = (status or "").upper()
    if table in ("ofc_links", "alternate_media"):
        return s in ("DOWN", "DEGRADED")
    if table == "terminal_equipment":
        return s in ("FAULTY", "EOL")
    if table == "radio_equipment":
        return s in ("FAULTY", "EOL")
    return False


def _status_severity(table, status):
    s = (status or "").upper()
    if s in ("DOWN", "FAULTY"):
        return "HIGH"
    if s in ("DEGRADED", "EOL"):
        return "MEDIUM"
    return "LOW"


def auto_restore_resource(conn, fault):
    """When a fault is marked RESOLVED, restore the referenced resource to its
    good state. Uses the fault's category + descriptive fields to find what
    needs restoring."""
    category = (fault["category"] or "").upper()
    desc = fault["description"] or ""
    affected_link = fault["affected_link"]
    affected_node = fault["affected_node"]

    if category == "OFC" and affected_link:
        conn.execute(
            "UPDATE ofc_links SET status='UP' WHERE link_id=? AND status != 'UP'",
            (affected_link,),
        )
    elif category == "ALT-MEDIA":
        # The description format is "<Media type> (<from>↔<to>) status changed to ..."
        # We can match on affected_node (from_node) and on description prefix.
        # Find the most recently-faulted alt media link matching this from_node.
        if affected_node:
            # Extract media type from description e.g. "Microwave (...)"
            media = desc.split(" ")[0] if desc else None
            if media:
                conn.execute(
                    """UPDATE alternate_media SET status='UP'
                       WHERE from_node=? AND media_type=? AND status != 'UP'""",
                    (affected_node, media),
                )
    elif category == "EQUIPMENT":
        # Extract equipment_id from description: "Terminal equipment <EQ-ID> status..."
        parts = desc.split(" ")
        if len(parts) >= 3:
            eq_id = parts[2]
            conn.execute(
                "UPDATE terminal_equipment SET status='OPERATIONAL' WHERE equipment_id=? AND status != 'OPERATIONAL'",
                (eq_id,),
            )
    elif category == "RADIO":
        # "Radio equipment <RAD-ID> (<type>) status..."
        parts = desc.split(" ")
        if len(parts) >= 3:
            rad_id = parts[2]
            conn.execute(
                "UPDATE radio_equipment SET status='OPERATIONAL' WHERE equipment_id=? AND status != 'OPERATIONAL'",
                (rad_id,),
            )
    elif category == "FIBER" and affected_link:
        # "Fiber <N> on link <LINK-ID> reported DOWN"
        parts = desc.split(" ")
        if len(parts) >= 2:
            try:
                fnum = int(parts[1])
                # Find the link, then update matching fibers
                link = conn.execute(
                    "SELECT id FROM ofc_links WHERE link_id=?", (affected_link,)
                ).fetchone()
                if link:
                    rib_ids = [r["id"] for r in conn.execute(
                        "SELECT id FROM ofc_ribbons WHERE ofc_link_id=?", (link["id"],)
                    ).fetchall()]
                    for rid in rib_ids:
                        # Restore the fiber: if it has both ports connected, CONNECTED-ACTIVE; else FREE
                        fiber = conn.execute(
                            "SELECT * FROM ofc_fibers WHERE ribbon_id=? AND fiber_number=?",
                            (rid, fnum),
                        ).fetchone()
                        if fiber and fiber["status"] == "CONNECTED-DOWN":
                            new_st = "CONNECTED-ACTIVE" if (fiber["from_port_id"] and fiber["to_port_id"]) else "FREE"
                            conn.execute(
                                "UPDATE ofc_fibers SET status=? WHERE id=?",
                                (new_st, fiber["id"]),
                            )
                            # Also clear the DOWN status on any equipment ports referenced
                            for pid in (fiber["from_port_id"], fiber["to_port_id"]):
                                if pid:
                                    conn.execute(
                                        "UPDATE equipment_ports SET status='ACTIVE' WHERE id=? AND status='DOWN'",
                                        (pid,),
                                    )
            except ValueError:
                pass
    elif category == "PORT":
        # "Port <label> on <EQ-ID> is DOWN"
        parts = desc.split(" ")
        if len(parts) >= 5:
            port_label = parts[1]
            eq_id = parts[3]
            eq = conn.execute(
                "SELECT id FROM terminal_equipment WHERE equipment_id=?", (eq_id,)
            ).fetchone()
            if eq:
                conn.execute(
                    """UPDATE equipment_ports
                       SET status=CASE WHEN status='DOWN' THEN 'ACTIVE' ELSE status END
                       WHERE terminal_equipment_id=? AND port_label=?""",
                    (eq["id"], port_label),
                )


@app.route("/api/ofc-detail/<int:ofc_id>")
def api_ofc_detail(ofc_id):
    """Return OFC link with ribbons, fibers, and resolved port info for both ends."""
    with closing(get_db()) as conn:
        link = conn.execute("SELECT * FROM ofc_links WHERE id=?", (ofc_id,)).fetchone()
        if not link:
            return jsonify({"error": "not found"}), 404
        ribbons = conn.execute(
            "SELECT * FROM ofc_ribbons WHERE ofc_link_id=? ORDER BY ribbon_number", (ofc_id,)
        ).fetchall()
        result_ribbons = []
        for rb in ribbons:
            fibers = conn.execute(
                """SELECT f.*,
                          fp.port_label AS from_port_label, fp.bandwidth AS from_port_bw,
                          ft.equipment_id AS from_equipment_id,
                          tp.port_label AS to_port_label, tp.bandwidth AS to_port_bw,
                          tt.equipment_id AS to_equipment_id
                   FROM ofc_fibers f
                   LEFT JOIN equipment_ports fp ON f.from_port_id = fp.id
                   LEFT JOIN terminal_equipment ft ON fp.terminal_equipment_id = ft.id
                   LEFT JOIN equipment_ports tp ON f.to_port_id = tp.id
                   LEFT JOIN terminal_equipment tt ON tp.terminal_equipment_id = tt.id
                   WHERE f.ribbon_id=? ORDER BY f.fiber_number""", (rb["id"],)
            ).fetchall()
            result_ribbons.append({**dict(rb), "fibers": rows_to_list(fibers)})
    return jsonify({**dict(link), "ribbons": result_ribbons})


@app.route("/api/equipment-detail/<int:eq_id>")
def api_equipment_detail(eq_id):
    """Return terminal equipment with all ports and the fiber they're connected to (if any)."""
    with closing(get_db()) as conn:
        eq = conn.execute("SELECT * FROM terminal_equipment WHERE id=?", (eq_id,)).fetchone()
        if not eq:
            return jsonify({"error": "not found"}), 404
        ports = conn.execute(
            """SELECT p.*,
                      f1.id AS fiber_id_from, f1.fiber_number AS fiber_num_from,
                      r1.ribbon_number AS ribbon_num_from, l1.link_id AS link_id_from,
                      f2.id AS fiber_id_to, f2.fiber_number AS fiber_num_to,
                      r2.ribbon_number AS ribbon_num_to, l2.link_id AS link_id_to
               FROM equipment_ports p
               LEFT JOIN ofc_fibers f1 ON f1.from_port_id = p.id
               LEFT JOIN ofc_ribbons r1 ON f1.ribbon_id = r1.id
               LEFT JOIN ofc_links l1 ON r1.ofc_link_id = l1.id
               LEFT JOIN ofc_fibers f2 ON f2.to_port_id = p.id
               LEFT JOIN ofc_ribbons r2 ON f2.ribbon_id = r2.id
               LEFT JOIN ofc_links l2 ON r2.ofc_link_id = l2.id
               WHERE p.terminal_equipment_id=?
               ORDER BY p.port_type, p.id""", (eq_id,)
        ).fetchall()
        port_list = []
        for p in ports:
            d = dict(p)
            # Combine the from/to fiber info into a single "connected_fiber" view
            if d["fiber_id_from"]:
                d["connected_fiber"] = {
                    "fiber_id": d["fiber_id_from"],
                    "fiber_number": d["fiber_num_from"],
                    "ribbon_number": d["ribbon_num_from"],
                    "link_id": d["link_id_from"],
                    "side": "FROM",
                }
            elif d["fiber_id_to"]:
                d["connected_fiber"] = {
                    "fiber_id": d["fiber_id_to"],
                    "fiber_number": d["fiber_num_to"],
                    "ribbon_number": d["ribbon_num_to"],
                    "link_id": d["link_id_to"],
                    "side": "TO",
                }
            else:
                d["connected_fiber"] = None
            port_list.append(d)
    return jsonify({**dict(eq), "ports": port_list})


@app.route("/api/available-ports/<node>")
def api_available_ports(node):
    """List terminal equipment + ports at a given node, used by the fiber-connect UI."""
    with closing(get_db()) as conn:
        rows = conn.execute(
            """SELECT p.id, p.port_label, p.port_type, p.bandwidth, p.status,
                      t.equipment_id, t.id AS terminal_id
               FROM equipment_ports p
               JOIN terminal_equipment t ON p.terminal_equipment_id = t.id
               WHERE t.location=?
               ORDER BY t.equipment_id, p.id""", (node,)
        ).fetchall()
    return jsonify(rows_to_list(rows))


@app.route("/api/fiber/<int:fiber_id>", methods=["PUT"])
def api_update_fiber(fiber_id):
    """Update a fiber's status and connected ports. Also keeps port statuses in sync."""
    data = request.get_json() or {}
    with closing(get_db()) as conn:
        old = conn.execute("SELECT * FROM ofc_fibers WHERE id=?", (fiber_id,)).fetchone()
        if not old:
            return jsonify({"error": "not found"}), 404
        new_status = data.get("status", old["status"])
        new_from = data.get("from_port_id", old["from_port_id"])
        new_to = data.get("to_port_id", old["to_port_id"])
        # Empty string -> NULL
        if new_from in ("", "null", "None"):
            new_from = None
        if new_to in ("", "null", "None"):
            new_to = None
        conn.execute(
            "UPDATE ofc_fibers SET status=?, from_port_id=?, to_port_id=?, remarks=? WHERE id=?",
            (new_status, new_from, new_to, data.get("remarks", old["remarks"]), fiber_id),
        )
        # Free up previously-used ports if they are no longer referenced anywhere
        for old_port in (old["from_port_id"], old["to_port_id"]):
            if old_port and old_port not in (new_from, new_to):
                still_used = conn.execute(
                    "SELECT 1 FROM ofc_fibers WHERE (from_port_id=? OR to_port_id=?) AND id!=?",
                    (old_port, old_port, fiber_id),
                ).fetchone()
                if not still_used:
                    conn.execute("UPDATE equipment_ports SET status='UNUSED' WHERE id=?", (old_port,))
        # Update new ports' status
        port_state = {
            "CONNECTED-ACTIVE": "ACTIVE",
            "CONNECTED-DOWN": "DOWN",
            "RESERVED": "ACTIVE",
            "FREE": "UNUSED",
        }.get(new_status, "UNUSED")
        for p in (new_from, new_to):
            if p:
                conn.execute("UPDATE equipment_ports SET status=? WHERE id=?", (port_state, p))

        # Auto-fault when a fiber transitions into CONNECTED-DOWN
        if (new_status == "CONNECTED-DOWN" and old["status"] != "CONNECTED-DOWN"):
            link_row = conn.execute(
                """SELECT l.link_id FROM ofc_links l
                   JOIN ofc_ribbons r ON r.ofc_link_id = l.id
                   WHERE r.id = ?""",
                (old["ribbon_id"],),
            ).fetchone()
            link_label = link_row["link_id"] if link_row else "?"
            auto_create_fault(conn,
                category="FIBER", severity="HIGH",
                description=f"Fiber {old['fiber_number']} on link {link_label} reported DOWN",
                affected_link=link_label)

        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/equipment-port", methods=["POST"])
def api_create_port():
    """Create a port on a terminal equipment (for the equipment-detail page)."""
    data = request.get_json() or {}
    required = ["terminal_equipment_id", "port_label"]
    for r in required:
        if not data.get(r):
            return jsonify({"error": f"{r} required"}), 400
    try:
        with closing(get_db()) as conn:
            cur = conn.execute(
                """INSERT INTO equipment_ports
                   (terminal_equipment_id, port_label, port_type, bandwidth, status, remarks)
                   VALUES (?,?,?,?,?,?)""",
                (data["terminal_equipment_id"], data["port_label"],
                 data.get("port_type", ""), data.get("bandwidth", ""),
                 data.get("status", "UNUSED"), data.get("remarks", "")),
            )
            conn.commit()
            return jsonify({"ok": True, "id": cur.lastrowid})
    except sqlite3.IntegrityError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/equipment-port/<int:port_id>", methods=["PUT", "DELETE"])
def api_modify_port(port_id):
    if request.method == "DELETE":
        with closing(get_db()) as conn:
            conn.execute("DELETE FROM equipment_ports WHERE id=?", (port_id,))
            conn.commit()
        return jsonify({"ok": True})
    data = request.get_json() or {}
    fields = ["port_label", "port_type", "bandwidth", "status", "remarks"]
    sets = ", ".join(f"{f}=?" for f in fields if f in data)
    if not sets:
        return jsonify({"error": "no fields"}), 400
    values = [data[f] for f in fields if f in data]
    values.append(port_id)
    with closing(get_db()) as conn:
        old = conn.execute("SELECT * FROM equipment_ports WHERE id=?", (port_id,)).fetchone()
        old_status = old["status"] if old else None
        conn.execute(f"UPDATE equipment_ports SET {sets} WHERE id=?", values)

        # Auto-fault if status transitioned into DOWN
        new_status = data.get("status")
        if new_status == "DOWN" and old_status != "DOWN" and old:
            eq = conn.execute(
                "SELECT equipment_id, location FROM terminal_equipment WHERE id=?",
                (old["terminal_equipment_id"],),
            ).fetchone()
            if eq:
                auto_create_fault(conn,
                    category="PORT", severity="HIGH",
                    description=f"Port {old['port_label']} on {eq['equipment_id']} is DOWN",
                    affected_node=eq["location"])

        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/ofcs-from-node/<node>")
def api_ofcs_from_node(node):
    """Return all OFC links touching this node, each with their ribbons and fibers
    (used by the port-side flow: 'connect this port to a fiber')."""
    with closing(get_db()) as conn:
        ofcs = conn.execute(
            "SELECT * FROM ofc_links WHERE from_node=? OR to_node=? ORDER BY link_id",
            (node, node),
        ).fetchall()
        result = []
        for ol in ofcs:
            ribbons = conn.execute(
                "SELECT * FROM ofc_ribbons WHERE ofc_link_id=? ORDER BY ribbon_number", (ol["id"],)
            ).fetchall()
            rb_list = []
            for rb in ribbons:
                fibers = conn.execute(
                    "SELECT id, fiber_number, status, from_port_id, to_port_id FROM ofc_fibers WHERE ribbon_id=? ORDER BY fiber_number",
                    (rb["id"],),
                ).fetchall()
                rb_list.append({**dict(rb), "fibers": rows_to_list(fibers)})
            result.append({**dict(ol), "ribbons": rb_list})
    return jsonify(result)


@app.route("/api/equipment-port/<int:port_id>", methods=["GET"])
def api_get_port(port_id):
    with closing(get_db()) as conn:
        p = conn.execute("SELECT * FROM equipment_ports WHERE id=?", (port_id,)).fetchone()
        if not p:
            return jsonify({"error": "not found"}), 404
        return jsonify(dict(p))


@app.route("/api/nodes")
@login_required
def api_nodes():
    extra, params = scope_filter("nodes")
    with closing(get_db()) as conn:
        sql = "SELECT * FROM nodes WHERE 1=1 " + extra + " ORDER BY name"
        rows = conn.execute(sql, params).fetchall()
    states = compute_node_states()
    out = []
    for r in rows:
        d = dict(r)
        d["state"] = states.get(d["name"], {}).get("state", "UNKNOWN")
        out.append(d)
    return jsonify(out)


@app.route("/api/ofc")
@login_required
def api_ofc():
    extra, params = scope_filter("ofc_links")
    with closing(get_db()) as conn:
        rows = conn.execute(
            "SELECT * FROM ofc_links WHERE 1=1 " + extra + " ORDER BY link_id", params
        ).fetchall()
    return jsonify(rows_to_list(rows))


@app.route("/api/alternate")
@login_required
def api_alternate():
    extra, params = scope_filter("alternate_media")
    with closing(get_db()) as conn:
        rows = conn.execute(
            "SELECT * FROM alternate_media WHERE 1=1 " + extra + " ORDER BY id", params
        ).fetchall()
    return jsonify(rows_to_list(rows))


@app.route("/api/equipment")
@login_required
def api_equipment():
    extra, params = scope_filter("terminal_equipment")
    with closing(get_db()) as conn:
        rows = conn.execute(
            "SELECT * FROM terminal_equipment WHERE 1=1 " + extra + " ORDER BY equipment_id", params
        ).fetchall()
    return jsonify(rows_to_list(rows))


@app.route("/api/radio")
@login_required
def api_radio():
    extra, params = scope_filter("radio_equipment")
    with closing(get_db()) as conn:
        rows = conn.execute(
            "SELECT * FROM radio_equipment WHERE 1=1 " + extra + " ORDER BY equipment_id", params
        ).fetchall()
    return jsonify(rows_to_list(rows))


@app.route("/api/faults")
@login_required
def api_faults():
    extra, params = scope_filter("fault_logs")
    with closing(get_db()) as conn:
        rows = conn.execute(
            "SELECT * FROM fault_logs WHERE 1=1 " + extra + " ORDER BY reported_at DESC", params
        ).fetchall()
    return jsonify(rows_to_list(rows))


@app.route("/api/node-detail/<name>")
@login_required
def api_node_detail(name):
    """Return all OFC links, alt media, terminal eq, radio eq for a single node."""
    with closing(get_db()) as conn:
        node = conn.execute("SELECT * FROM nodes WHERE name=?", (name,)).fetchone()
        if not node:
            return jsonify({"error": "node not found"}), 404
        ofc = conn.execute(
            "SELECT * FROM ofc_links WHERE from_node=? OR to_node=? ORDER BY status DESC, link_id",
            (name, name),
        ).fetchall()
        alt = conn.execute(
            "SELECT * FROM alternate_media WHERE from_node=? OR to_node=? ORDER BY status DESC, media_type",
            (name, name),
        ).fetchall()
        eq = conn.execute(
            "SELECT * FROM terminal_equipment WHERE location=? ORDER BY equipment_id",
            (name,),
        ).fetchall()
        rad = conn.execute(
            "SELECT * FROM radio_equipment WHERE location=? ORDER BY equipment_id",
            (name,),
        ).fetchall()
        states = compute_node_states()
    return jsonify({
        "node": dict(node),
        "state": states.get(name, {}).get("state", "UNKNOWN"),
        "ofc_links": rows_to_list(ofc),
        "alt_media": rows_to_list(alt),
        "terminal_equipment": rows_to_list(eq),
        "radio_equipment": rows_to_list(rad),
    })


@app.route("/api/kpis")
@login_required
def api_kpis():
    return jsonify(compute_kpis())


@app.route("/api/node-states")
@login_required
def api_node_states():
    return jsonify(compute_node_states())


@app.route("/api/analytics-data")
@login_required
def api_analytics_data():
    flt_extra, flt_p = scope_filter("fault_logs")
    eq_extra, eq_p = scope_filter("terminal_equipment")
    ofc_extra, ofc_p = scope_filter("ofc_links")
    with closing(get_db()) as conn:
        # Faults by category
        cat = conn.execute(
            "SELECT category, COUNT(*) c FROM fault_logs WHERE 1=1 " + flt_extra +
            " GROUP BY category", flt_p
        ).fetchall()
        # Faults by severity (open only)
        sev = conn.execute(
            "SELECT severity, COUNT(*) c FROM fault_logs WHERE status='OPEN' " + flt_extra +
            " GROUP BY severity", flt_p
        ).fetchall()
        # Equipment age distribution
        eq = conn.execute(
            "SELECT year_purchased, COUNT(*) c FROM terminal_equipment WHERE 1=1 " + eq_extra +
            " GROUP BY year_purchased ORDER BY year_purchased", eq_p
        ).fetchall()
        # OFC age distribution
        ofc_age = conn.execute(
            "SELECT year_laid, COUNT(*) c FROM ofc_links WHERE 1=1 " + ofc_extra +
            " GROUP BY year_laid ORDER BY year_laid", ofc_p
        ).fetchall()
        # Loss vs margin
        ofc_loss = conn.execute(
            "SELECT link_id, loss_db, margin_db, status FROM ofc_links WHERE 1=1 " + ofc_extra, ofc_p
        ).fetchall()
        # Faults timeline (last 30 days)
        thirty = (datetime.now() - timedelta(days=30)).isoformat()
        timeline = conn.execute(
            "SELECT date(reported_at) d, COUNT(*) c FROM fault_logs "
            "WHERE reported_at >= ? " + flt_extra +
            " GROUP BY date(reported_at) ORDER BY d",
            (thirty, *flt_p)
        ).fetchall()

    return jsonify({
        "faults_by_category": rows_to_list(cat),
        "open_faults_by_severity": rows_to_list(sev),
        "equipment_by_year": rows_to_list(eq),
        "ofc_by_year": rows_to_list(ofc_age),
        "ofc_loss_margin": rows_to_list(ofc_loss),
        "faults_timeline": rows_to_list(timeline),
        "mttr": mttr_stats(),
    })


def _categorize_fault(category):
    """Map a fault's category to one of the 4 display buckets."""
    c = (category or "").upper()
    if c in ("OFC", "FIBER"):
        return "Communication Links"
    if c == "ALT-MEDIA":
        return "Alternate Media"
    if c in ("EQUIPMENT", "PORT"):
        return "Terminal Equipment"
    if c == "RADIO":
        return "Radio Equipment"
    return "Other"


BUCKETS = ["Communication Links", "Alternate Media", "Terminal Equipment", "Radio Equipment"]


def _node_formation_map(conn):
    """name -> formation lookup."""
    return {r["name"]: r["formation"]
            for r in conn.execute("SELECT name, formation FROM nodes").fetchall()}


def _fault_formation(fault, node_form, conn):
    """Resolve which formation a fault belongs to via its affected node/link."""
    node = fault["affected_node"]
    if node and node in node_form and node_form[node]:
        return node_form[node]
    link = fault["affected_link"]
    if link:
        r = conn.execute(
            "SELECT from_node, to_node FROM ofc_links WHERE link_id=?", (link,)
        ).fetchone()
        if r:
            for n in (r["from_node"], r["to_node"]):
                if n in node_form and node_form[n]:
                    return node_form[n]
    return "(Unassigned)"


def _group_faults_by_formation(rows, conn):
    """Return {formation: {bucket: count, ...}, ...} plus per-formation totals."""
    node_form = _node_formation_map(conn)
    result = {}
    for f in rows:
        formation = _fault_formation(f, node_form, conn)
        bucket = _categorize_fault(f["category"])
        if bucket not in BUCKETS:
            continue
        result.setdefault(formation, {b: 0 for b in BUCKETS})
        result[formation][bucket] += 1
    return result


@app.route("/api/daily-outage-data")
@login_required
def api_daily_outage_data():
    """Returns today's fresh faults and all-time faults, both grouped by
    formation and sub-categorized into the 4 buckets."""
    flt_extra, flt_p = scope_filter("fault_logs")
    today = date.today().isoformat()
    with closing(get_db()) as conn:
        today_rows = conn.execute(
            "SELECT * FROM fault_logs WHERE date(reported_at)=? " + flt_extra,
            (today, *flt_p)
        ).fetchall()
        all_rows = conn.execute(
            "SELECT * FROM fault_logs WHERE 1=1 " + flt_extra, flt_p
        ).fetchall()

        today_grouped = _group_faults_by_formation(today_rows, conn)
        total_grouped = _group_faults_by_formation(all_rows, conn)

    return jsonify({
        "buckets": BUCKETS,
        "today": today_grouped,
        "total": total_grouped,
        "today_count": len(today_rows),
        "total_count": len(all_rows),
        "date": today,
    })


@app.route("/api/misc-activities", methods=["GET"])
@login_required
def api_list_misc_activities():
    """List misc activities with role-based visibility:
       - Admin & Div roles see ALL activities
       - Other (Bde etc.) roles see only their own authored activities."""
    u = current_user()
    role = u["role"]
    is_div_or_admin = (role == "Admin") or ("div" in role.lower())
    with closing(get_db()) as conn:
        if is_div_or_admin:
            rows = conn.execute(
                "SELECT * FROM misc_activities ORDER BY created_at DESC"
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM misc_activities WHERE author_username=? ORDER BY created_at DESC",
                (u["username"],)
            ).fetchall()
    out = rows_to_list(rows)
    # Tell the frontend whether the current user can edit/delete each row
    for r in out:
        r["can_modify"] = (r["author_username"] == u["username"]) or (role == "Admin")
    return jsonify({"activities": out, "viewer_role": role, "viewer": u["username"]})


@app.route("/api/misc-activities", methods=["POST"])
@login_required
def api_create_misc_activity():
    u = current_user()
    data = request.get_json() or {}
    content = (data.get("content") or "").strip()
    if not content:
        return jsonify({"error": "content required"}), 400
    now = datetime.now().isoformat(timespec="seconds")
    with closing(get_db()) as conn:
        conn.execute(
            """INSERT INTO misc_activities
               (content, author_username, author_role, activity_date, created_at, updated_at)
               VALUES (?,?,?,?,?,?)""",
            (content, u["username"], u["role"], date.today().isoformat(), now, now),
        )
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/misc-activities/<int:act_id>", methods=["PUT", "DELETE"])
@login_required
def api_modify_misc_activity(act_id):
    u = current_user()
    with closing(get_db()) as conn:
        row = conn.execute("SELECT * FROM misc_activities WHERE id=?", (act_id,)).fetchone()
        if not row:
            return jsonify({"error": "not found"}), 404
        # Only the author or an Admin may edit/delete
        if row["author_username"] != u["username"] and u["role"] != "Admin":
            return jsonify({"error": "not authorized to modify this entry"}), 403
        if request.method == "DELETE":
            conn.execute("DELETE FROM misc_activities WHERE id=?", (act_id,))
            conn.commit()
            return jsonify({"ok": True})
        data = request.get_json() or {}
        content = (data.get("content") or "").strip()
        if not content:
            return jsonify({"error": "content required"}), 400
        conn.execute(
            "UPDATE misc_activities SET content=?, updated_at=? WHERE id=?",
            (content, datetime.now().isoformat(timespec="seconds"), act_id),
        )
        conn.commit()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# API routes - CRUD (manual entry through UI)
# ---------------------------------------------------------------------------
TABLE_FIELDS = {
    "nodes": ["name", "latitude", "longitude", "region", "node_type",
              "formation", "unit", "coy", "contact_no", "remarks"],
    "ofc_links": ["link_id", "from_node", "to_node", "distance_km", "year_laid",
                  "no_of_fiber", "ribbon_count", "loss_db", "no_dark_fiber", "cable_type",
                  "margin_db", "status", "last_trace_date", "trace_taken_by", "remarks"],
    "alternate_media": ["media_type", "from_node", "to_node", "spec",
                        "hop_distance_km", "status", "remarks"],
    "terminal_equipment": ["equipment_id", "location", "eqpt_type", "eth_ports",
                           "e1_voice_ports", "capacity", "year_purchased",
                           "status", "last_checked_on", "last_checked_by", "remarks"],
    "radio_equipment": ["equipment_id", "radio_type", "location", "frequency",
                        "year_purchased", "status", "remarks"],
    "fault_logs": ["fault_id", "category", "affected_link", "affected_node",
                   "reported_at", "resolved_at", "severity", "description",
                   "action_taken", "status"],
}


@app.route("/api/<table>/<int:row_id>", methods=["PUT", "DELETE"])
def crud_row(table, row_id):
    if table not in TABLE_FIELDS:
        return jsonify({"error": "unknown table"}), 400

    if request.method == "DELETE":
        with closing(get_db()) as conn:
            conn.execute(f"DELETE FROM {table} WHERE id=?", (row_id,))
            conn.commit()
        return jsonify({"ok": True})

    data = request.get_json() or {}
    fields = TABLE_FIELDS[table]
    sets = ", ".join(f"{f}=?" for f in fields if f in data)
    if not sets:
        return jsonify({"error": "no fields"}), 400
    values = [data[f] for f in fields if f in data]
    if "updated_at" in [c[1] for c in get_db().execute(f"PRAGMA table_info({table})").fetchall()]:
        sets += ", updated_at=?"
        values.append(datetime.now().isoformat(timespec="seconds"))
    values.append(row_id)
    with closing(get_db()) as conn:
        # Read existing row first so we can compare status transitions for auto-fault
        existing = conn.execute(f"SELECT * FROM {table} WHERE id=?", (row_id,)).fetchone()
        old_status = existing["status"] if existing and "status" in existing.keys() else None

        conn.execute(f"UPDATE {table} SET {sets} WHERE id=?", values)

        # Auto-fault: if status changed into a "bad" state, log a fault
        new_status = data.get("status")
        if (new_status and new_status != old_status
                and _status_is_bad(table, new_status)):
            sev = _status_severity(table, new_status)
            updated = conn.execute(f"SELECT * FROM {table} WHERE id=?", (row_id,)).fetchone()
            if table == "ofc_links":
                auto_create_fault(conn,
                    category="OFC", severity=sev,
                    description=f"OFC link {updated['link_id']} status changed to {new_status}",
                    affected_link=updated["link_id"])
            elif table == "alternate_media":
                desc_node = updated["from_node"] if not updated["to_node"] else f"{updated['from_node']}↔{updated['to_node']}"
                auto_create_fault(conn,
                    category="ALT-MEDIA", severity=sev,
                    description=f"{updated['media_type']} ({desc_node}) status changed to {new_status}",
                    affected_node=updated["from_node"])
            elif table == "terminal_equipment":
                auto_create_fault(conn,
                    category="EQUIPMENT", severity=sev,
                    description=f"Terminal equipment {updated['equipment_id']} status changed to {new_status}",
                    affected_node=updated["location"])
            elif table == "radio_equipment":
                auto_create_fault(conn,
                    category="RADIO", severity=sev,
                    description=f"Radio equipment {updated['equipment_id']} ({updated['radio_type']}) status changed to {new_status}",
                    affected_node=updated["location"])

        # Auto-restore: when a fault is marked RESOLVED, restore the underlying resource
        if (table == "fault_logs" and new_status == "RESOLVED"
                and old_status != "RESOLVED" and existing):
            auto_restore_resource(conn, existing)
            # Also stamp resolved_at if not present in this update
            if "resolved_at" not in data:
                conn.execute(
                    "UPDATE fault_logs SET resolved_at=? WHERE id=?",
                    (datetime.now().isoformat(timespec="seconds"), row_id),
                )

        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/<table>", methods=["POST"])
def crud_create(table):
    if table not in TABLE_FIELDS:
        return jsonify({"error": "unknown table"}), 400
    data = request.get_json() or {}
    fields = TABLE_FIELDS[table]
    cols = [f for f in fields if f in data]
    if not cols:
        return jsonify({"error": "no fields"}), 400
    placeholders = ", ".join("?" for _ in cols)
    cols_sql = ", ".join(cols)
    values = [data[f] for f in cols]
    has_updated = "updated_at" in [c[1] for c in get_db().execute(f"PRAGMA table_info({table})").fetchall()]
    if has_updated:
        cols_sql += ", updated_at"
        placeholders += ", ?"
        values.append(datetime.now().isoformat(timespec="seconds"))
    try:
        with closing(get_db()) as conn:
            cur = conn.execute(
                f"INSERT INTO {table} ({cols_sql}) VALUES ({placeholders})", values
            )
            new_id = cur.lastrowid
            # Auto-generate ribbons + fibers when a new OFC link is created
            if table == "ofc_links":
                n_fibers = int(data.get("no_of_fiber") or 0)
                n_ribbons = max(1, int(data.get("ribbon_count") or 1))
                if n_fibers > 0:
                    base = n_fibers // n_ribbons
                    extra = n_fibers % n_ribbons
                    for rb in range(1, n_ribbons + 1):
                        cnt = base + (1 if rb <= extra else 0)
                        rcur = conn.execute(
                            "INSERT INTO ofc_ribbons (ofc_link_id, ribbon_number, fiber_count) VALUES (?,?,?)",
                            (new_id, rb, cnt),
                        )
                        rid = rcur.lastrowid
                        for fnum in range(1, cnt + 1):
                            conn.execute(
                                "INSERT INTO ofc_fibers (ribbon_id, fiber_number, status) VALUES (?,?,?)",
                                (rid, fnum, "FREE"),
                            )
            # Auto-generate ports when a new terminal equipment is created.
            # Look up the equipment type template; if found, use it.
            elif table == "terminal_equipment":
                eqpt_type_name = data.get("eqpt_type")
                template = []
                if eqpt_type_name:
                    et_row = conn.execute(
                        "SELECT port_template FROM equipment_types WHERE name=?",
                        (eqpt_type_name,),
                    ).fetchone()
                    if et_row and et_row["port_template"]:
                        try:
                            template = json.loads(et_row["port_template"])
                        except Exception:
                            template = []

                if template:
                    # Generate ports from template
                    for spec in template:
                        prefix = spec.get("label_prefix", "Port-")
                        ptype = spec.get("port_type", "Optical")
                        bw = spec.get("bandwidth", "1 Gbps")
                        cnt = int(spec.get("count", 0) or 0)
                        for p in range(1, cnt + 1):
                            try:
                                conn.execute(
                                    "INSERT INTO equipment_ports (terminal_equipment_id, port_label, port_type, bandwidth, status) VALUES (?,?,?,?,?)",
                                    (new_id, f"{prefix}{p}", ptype, bw, "UNUSED"),
                                )
                            except Exception:
                                pass
                else:
                    # Fallback to legacy eth/e1 fields if no template found
                    eth = int(data.get("eth_ports") or 0)
                    e1 = int(data.get("e1_voice_ports") or 0)
                    base_bw = data.get("capacity") or "1 Gbps"
                    for p in range(1, eth + 1):
                        bw = base_bw if p <= max(1, eth // 4) else "1 Gbps"
                        conn.execute(
                            "INSERT INTO equipment_ports (terminal_equipment_id, port_label, port_type, bandwidth, status) VALUES (?,?,?,?,?)",
                            (new_id, f"Eth-{p}", "Optical", bw, "UNUSED"),
                        )
                    for p in range(1, e1 + 1):
                        conn.execute(
                            "INSERT INTO equipment_ports (terminal_equipment_id, port_label, port_type, bandwidth, status) VALUES (?,?,?,?,?)",
                            (new_id, f"E1-{p}", "Electronic", "2 Mbps", "UNUSED"),
                        )
            conn.commit()
            return jsonify({"ok": True, "id": new_id})
    except sqlite3.IntegrityError as e:
        return jsonify({"error": str(e)}), 400


# ---------------------------------------------------------------------------
# Excel import / export
# ---------------------------------------------------------------------------
EXPECTED_SHEETS = {
    "Nodes": ("nodes", ["name", "latitude", "longitude", "region", "node_type",
                        "formation", "unit", "coy", "contact_no", "remarks"]),
    "OFC_Links": ("ofc_links", ["link_id", "from_node", "to_node", "distance_km",
                                "year_laid", "no_of_fiber", "ribbon_count", "loss_db", "no_dark_fiber",
                                "cable_type", "margin_db", "status",
                                "last_trace_date", "trace_taken_by", "remarks"]),
    "Alternate_Media": ("alternate_media", ["media_type", "from_node", "to_node",
                                            "spec", "hop_distance_km", "status", "remarks"]),
    "Terminal_Equipment": ("terminal_equipment", ["equipment_id", "location", "eqpt_type",
                                                   "eth_ports", "e1_voice_ports", "capacity",
                                                   "year_purchased", "status",
                                                   "last_checked_on", "last_checked_by", "remarks"]),
    "Radio_Equipment": ("radio_equipment", ["equipment_id", "radio_type", "location",
                                             "frequency", "year_purchased", "status", "remarks"]),
    "Equipment_Types": ("equipment_types", ["name", "description", "port_template", "remarks"]),
    "Fault_Logs": ("fault_logs", ["fault_id", "category", "affected_link", "affected_node",
                                   "reported_at", "resolved_at", "severity", "description",
                                   "action_taken", "status"]),
}

# Unique-key columns for sync mode: lets us match excel rows against existing DB rows
SHEET_KEY_COLS = {
    "Nodes": ["name"],
    "OFC_Links": ["link_id"],
    "Alternate_Media": ["media_type", "from_node", "to_node"],
    "Terminal_Equipment": ["equipment_id"],
    "Radio_Equipment": ["equipment_id"],
    "Equipment_Types": ["name"],
    "Fault_Logs": ["fault_id"],
}

# Sheets that need custom logic because they reference parent rows by composite keys
SPECIAL_SHEETS = {
    "Equipment_Ports": [
        "equipment_id",      # parent terminal_equipment lookup key
        "port_label",        # unique within equipment
        "port_type",         # Optical | Electronic
        "bandwidth",
        "status",
        "remarks",
    ],
    "OFC_Fibers": [
        "link_id",           # parent OFC link
        "ribbon_number",     # which ribbon
        "fiber_number",      # which fiber within the ribbon
        "status",            # FREE | CONNECTED-ACTIVE | CONNECTED-DOWN | RESERVED
        "from_equipment_id", # which equipment hosts the from-port (optional)
        "from_port_label",   # which port on it (optional)
        "to_equipment_id",
        "to_port_label",
        "remarks",
    ],
}


@app.route("/api/upload-excel", methods=["POST"])
@admin_required
def upload_excel():
    """Modes:
       - replace: wipe each table, then insert (legacy)
       - append : insert-or-ignore (legacy)
       - sync   : ADD new rows from excel, UPDATE existing rows that match by
                  unique key, and DELETE DB rows whose key is not present in
                  the excel sheet.  Only sheets actually included in the file
                  are affected (a sheet missing from the upload is left alone).
    """
    f = request.files.get("file")
    mode = request.form.get("mode", "sync")
    if mode not in ("replace", "append", "sync"):
        return jsonify({"error": f"Unknown mode '{mode}'"}), 400
    if not f:
        return jsonify({"error": "no file"}), 400
    try:
        sheets = pd.read_excel(f, sheet_name=None)
    except Exception as e:
        return jsonify({"error": f"Could not read Excel: {e}"}), 400

    summary = {}
    with closing(get_db()) as conn:
        for sheet_name, (table, cols) in EXPECTED_SHEETS.items():
            if sheet_name not in sheets:
                summary[sheet_name] = {"status": "skipped (not in file)"}
                continue
            df = sheets[sheet_name]
            df = df.where(pd.notna(df), None)
            df.columns = [str(c).strip() for c in df.columns]
            usable_cols = [c for c in cols if c in df.columns]
            if not usable_cols:
                summary[sheet_name] = {"status": "skipped (no recognized columns)"}
                continue

            has_updated = "updated_at" in [c[1] for c in
                conn.execute(f"PRAGMA table_info({table})").fetchall()]
            now_ts = datetime.now().isoformat(timespec="seconds")

            if mode == "sync":
                key_cols = SHEET_KEY_COLS.get(sheet_name)
                if not key_cols or any(k not in usable_cols for k in key_cols):
                    # Fallback to append if we can't sync
                    summary[sheet_name] = {
                        "status": f"sync unsupported (missing key cols), falling back to append"}
                    mode_for_sheet = "append"
                else:
                    mode_for_sheet = "sync"
            else:
                mode_for_sheet = mode

            if mode_for_sheet == "replace":
                conn.execute(f"DELETE FROM {table}")
                inserted, updated, deleted = 0, 0, 0
                for _, row in df.iterrows():
                    vals = [row.get(c) for c in usable_cols]
                    cs = list(usable_cols)
                    if has_updated:
                        cs.append("updated_at"); vals.append(now_ts)
                    placeholders = ", ".join("?" for _ in cs)
                    try:
                        conn.execute(
                            f"INSERT INTO {table} ({', '.join(cs)}) VALUES ({placeholders})",
                            vals)
                        inserted += 1
                    except Exception:
                        continue
                summary[sheet_name] = {
                    "status": "ok", "mode": "replace",
                    "inserted": inserted, "updated": 0, "deleted": 0}
                continue

            if mode_for_sheet == "append":
                inserted = 0
                for _, row in df.iterrows():
                    vals = [row.get(c) for c in usable_cols]
                    cs = list(usable_cols)
                    if has_updated:
                        cs.append("updated_at"); vals.append(now_ts)
                    placeholders = ", ".join("?" for _ in cs)
                    try:
                        conn.execute(
                            f"INSERT OR IGNORE INTO {table} ({', '.join(cs)}) VALUES ({placeholders})",
                            vals)
                        inserted += 1
                    except Exception:
                        continue
                summary[sheet_name] = {
                    "status": "ok", "mode": "append",
                    "inserted": inserted, "updated": 0, "deleted": 0}
                continue

            # mode_for_sheet == "sync"
            key_cols = SHEET_KEY_COLS[sheet_name]
            non_key_cols = [c for c in usable_cols if c not in key_cols]

            # Existing DB rows keyed by tuple of key_cols
            existing = {}
            for r in conn.execute(
                f"SELECT id, {', '.join(key_cols)} FROM {table}").fetchall():
                k = tuple((r[c] if r[c] is not None else "") for c in key_cols)
                existing[k] = r["id"]

            inserted = updated = deleted = 0
            excel_keys = set()
            for _, row in df.iterrows():
                key_vals = tuple((row.get(c) if row.get(c) is not None else "") for c in key_cols)
                # Skip rows where all key cols are blank
                if all((v == "" or v is None) for v in key_vals):
                    continue
                excel_keys.add(key_vals)
                if key_vals in existing:
                    # UPDATE
                    if non_key_cols:
                        sets = ", ".join(f"{c}=?" for c in non_key_cols)
                        vals = [row.get(c) for c in non_key_cols]
                        if has_updated:
                            sets += ", updated_at=?"
                            vals.append(now_ts)
                        vals.append(existing[key_vals])
                        try:
                            conn.execute(
                                f"UPDATE {table} SET {sets} WHERE id=?", vals)
                            updated += 1
                        except Exception:
                            continue
                else:
                    # INSERT
                    vals = [row.get(c) for c in usable_cols]
                    cs = list(usable_cols)
                    if has_updated:
                        cs.append("updated_at"); vals.append(now_ts)
                    placeholders = ", ".join("?" for _ in cs)
                    try:
                        conn.execute(
                            f"INSERT INTO {table} ({', '.join(cs)}) VALUES ({placeholders})",
                            vals)
                        inserted += 1
                    except Exception:
                        continue

            # DELETE rows whose keys are not in the excel keyset
            for k, db_id in existing.items():
                if k not in excel_keys:
                    try:
                        conn.execute(f"DELETE FROM {table} WHERE id=?", (db_id,))
                        deleted += 1
                    except Exception:
                        continue

            summary[sheet_name] = {
                "status": "ok", "mode": "sync",
                "inserted": inserted, "updated": updated, "deleted": deleted}

        conn.commit()

    return jsonify({"ok": True, "summary": summary, "mode": mode})


@app.route("/api/export-excel")
@admin_required
def export_excel():
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        with closing(get_db()) as conn:
            for sheet_name, (table, cols) in EXPECTED_SHEETS.items():
                df = pd.read_sql_query(f"SELECT {', '.join(cols)} FROM {table}", conn)
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Equipment_Ports — joined with terminal_equipment to get equipment_id
            ports_df = pd.read_sql_query(
                """SELECT t.equipment_id AS equipment_id,
                          p.port_label, p.port_type, p.bandwidth, p.status, p.remarks
                   FROM equipment_ports p
                   JOIN terminal_equipment t ON p.terminal_equipment_id = t.id
                   ORDER BY t.equipment_id, p.id""", conn)
            ports_df.to_excel(writer, sheet_name="Equipment_Ports", index=False)

            # OFC_Fibers — joined with ribbons & links + resolved from/to ports
            fibers_df = pd.read_sql_query(
                """SELECT l.link_id AS link_id,
                          r.ribbon_number AS ribbon_number,
                          f.fiber_number AS fiber_number,
                          f.status AS status,
                          ft.equipment_id AS from_equipment_id,
                          fp.port_label  AS from_port_label,
                          tt.equipment_id AS to_equipment_id,
                          tp.port_label  AS to_port_label,
                          f.remarks AS remarks
                   FROM ofc_fibers f
                   JOIN ofc_ribbons r ON f.ribbon_id = r.id
                   JOIN ofc_links l ON r.ofc_link_id = l.id
                   LEFT JOIN equipment_ports fp ON f.from_port_id = fp.id
                   LEFT JOIN terminal_equipment ft ON fp.terminal_equipment_id = ft.id
                   LEFT JOIN equipment_ports tp ON f.to_port_id = tp.id
                   LEFT JOIN terminal_equipment tt ON tp.terminal_equipment_id = tt.id
                   ORDER BY l.link_id, r.ribbon_number, f.fiber_number""", conn)
            fibers_df.to_excel(writer, sheet_name="OFC_Fibers", index=False)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"wan_data_{date.today().isoformat()}.xlsx",
    )


@app.route("/api/template-excel")
@admin_required
def template_excel():
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, (_, cols) in EXPECTED_SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        for i, c in enumerate(cols, 1):
            ws.cell(row=1, column=i, value=c)
    # Special sheets
    for sheet_name, cols in SPECIAL_SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        for i, c in enumerate(cols, 1):
            ws.cell(row=1, column=i, value=c)
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="wan_template.xlsx",
    )


@app.route("/api/reset-mock", methods=["POST"])
@admin_required
def reset_mock():
    with closing(get_db()) as conn:
        for table in EXPECTED_SHEETS.values():
            conn.execute(f"DELETE FROM {table[0]}")
        conn.execute("DELETE FROM equipment_types")
        conn.commit()
    seed_default_equipment_types()
    seed_mock_data()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Bootstrap
# ---------------------------------------------------------------------------
with app.app_context():
    init_db()
    seed_default_equipment_types()
    seed_default_admin()


if __name__ == "__main__":
    print("=" * 60)
    print("Vajr Comn State")
    print("=" * 60)
    print(f"DB: {DB_PATH}")
    print("Default login: admin / admin  (change after first login)")
    print("Open: http://localhost:5000")
    print("=" * 60)
    app.run(host="0.0.0.0", port=5000, debug=False)
